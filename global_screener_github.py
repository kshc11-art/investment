#!/usr/bin/env python3
"""
글로벌 주식/ETF 스크리닝 - GitHub Actions 버전 v2.5
- v2.5 수정:
  * 한국 ETF: KRX 정보데이터시스템 직접 크롤링 (가장 정확)
  * KRX → pykrx → 네이버 → 폴백 순서로 시도
- v2.4 수정: pykrx.get_etf_ticker_list() 사용
- v2.3 수정: KOSPI 150개, 외국인 20일, ISM PMI 등

GitHub Actions에서 자동 실행 → JSON 출력 → GitHub Pages에서 PWA가 fetch

설치:
pip install yfinance openpyxl pandas requests beautifulsoup4 lxml numpy pykrx

실행:
python global_screener_github.py              # Excel + JSON 출력
python global_screener_github.py --json-only  # JSON만 출력
"""

import warnings
warnings.filterwarnings('ignore')

import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import time
import sys
import re
import json
import argparse

# ★ GitHub Actions 버퍼링 해결: 즉시 출력 함수
def log(msg):
    """즉시 출력되는 로그 함수"""
    print(msg)
    sys.stdout.flush()

log("라이브러리 로딩 중...")

import logging
logging.getLogger('yfinance').setLevel(logging.CRITICAL)
logging.getLogger('urllib3').setLevel(logging.CRITICAL)

try:
    import yfinance as yf
    import requests
    from bs4 import BeautifulSoup
except ImportError as e:
    log("=" * 60)
    log("pip install yfinance openpyxl pandas requests beautifulsoup4 lxml numpy pykrx")
    log("=" * 60)
    sys.exit(1)

# pykrx 선택적
try:
    from pykrx import stock as pykrx_stock
    PYKRX_AVAILABLE = True
    log("✅ pykrx 로드 완료")
except:
    PYKRX_AVAILABLE = False
    log("⚠️ pykrx 미설치 - 한국 주식 일부 데이터 제한")

# ============================================================
# 설정
# ============================================================
TODAY = datetime.now().strftime("%Y%m%d")
DATE_1Y_AGO = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")

# 종목 수 설정
TOP_N_KR = 150    # 한국 주식: KOSPI 150개 고정
TOP_N_US = 100    # 미국 주식: 100개
TOP_N_KR_ETF = 200  # 한국 ETF: 200개
TOP_N_US_ETF = 100  # 미국 ETF: 100개

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
}

# 네이버 차단 여부 (런타임에 판단)
NAVER_AVAILABLE = None
FNGUIDE_AVAILABLE = None  # ★ FnGuide도 차단 감지

# 타임아웃 설정 (GitHub Actions용 단축)
TIMEOUT_SHORT = 3
TIMEOUT_LONG = 5

# 데이터 품질 검증용 필수 컬럼
REQUIRED_COLS_KR_STOCK = ['PER', 'PBR', 'ROE(%)', 'Return60D(%)', 'RSI14', 'Volatility20D']
REQUIRED_COLS_US_STOCK = ['PER', 'PBR', 'ROE(%)', 'Return60D(%)', 'RSI14', 'Volatility20D', 'Beta']
REQUIRED_COLS_ETF = ['ExpenseRatio(%)', 'Return3M(%)', 'Volatility20D']

# ============================================================
# 유틸리티 함수
# ============================================================
def fmt(val, decimals=2):
    """숫자 포맷팅"""
    if val is None or pd.isna(val):
        return None
    try:
        return round(float(val), decimals)
    except:
        return None

def safe_get(d, *keys, default=None):
    """딕셔너리에서 안전하게 값 추출"""
    if not isinstance(d, dict):
        return default
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
    return default

def fetch_yf_with_retry(ticker, max_retries=2, delay=0.3):
    """yfinance 데이터를 재시도 로직과 함께 가져오기"""
    for attempt in range(max_retries):
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="1y")  # 2y → 1y로 단축
            info = t.info

            # 유효한 데이터인지 확인
            if not hist.empty or (info and len(info) > 5):
                return t, hist, info

            if attempt < max_retries - 1:
                time.sleep(delay)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(delay)
            else:
                pass

    return None, pd.DataFrame(), {}

def try_multiple_tickers(ticker_variants, max_retries=2):
    """여러 티커 형식을 시도하여 가장 좋은 결과 반환"""
    best_result = (None, pd.DataFrame(), {})
    best_score = 0

    for ticker in ticker_variants:
        t, hist, info = fetch_yf_with_retry(ticker, max_retries=max_retries)

        # 점수 계산: hist 길이 + info 키 수
        score = len(hist) + len(info) if info else len(hist)

        if score > best_score:
            best_score = score
            best_result = (t, hist, info)

            # 충분히 좋은 결과면 조기 종료
            if len(hist) > 200 and len(info) > 20:
                break

    return best_result

def fill_missing_from_info(row, info, field_mapping):
    """info 딕셔너리에서 결측값 채우기"""
    for row_field, (info_keys, multiplier, decimals) in field_mapping.items():
        if row.get(row_field) is None:
            val = safe_get(info, *info_keys) if isinstance(info_keys, tuple) else safe_get(info, info_keys)
            if val is not None:
                if multiplier != 1:
                    val = val * multiplier
                row[row_field] = fmt(val, decimals)
    return row

# ============================================================
# 네이버 금융 스크래핑
# ============================================================
def fetch_naver(url, timeout=TIMEOUT_SHORT):
    """네이버 URL 요청"""
    global NAVER_AVAILABLE
    if NAVER_AVAILABLE is False:
        return None
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code == 403:
            NAVER_AVAILABLE = False
            log("  ⚠️ 네이버 금융 접근 차단됨 (403)")
            return None
        if resp.status_code == 200:
            NAVER_AVAILABLE = True
            return resp
        return None
    except:
        return None

def get_naver_stock_list(market='KOSPI', max_pages=10):
    """네이버에서 시가총액 순위 종목 리스트"""
    global NAVER_AVAILABLE
    stocks = []
    sosok = '0' if market == 'KOSPI' else '1'
    
    log(f"  네이버 {market} 시총 순위 로드 시도...")
    
    for page in range(1, max_pages + 1):
        url = f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page={page}"
        resp = fetch_naver(url)
        
        if not resp:
            if NAVER_AVAILABLE is False:
                return []
            continue
        
        soup = BeautifulSoup(resp.text, 'lxml')
        table = soup.select_one('table.type_2')
        
        if not table:
            continue
        
        for tr in table.select('tr'):
            tds = tr.select('td')
            if len(tds) < 10:
                continue
            
            try:
                link = tds[1].select_one('a')
                if not link:
                    continue
                
                href = link.get('href', '')
                code_match = re.search(r'code=(\d{6})', href)
                if not code_match:
                    continue
                
                code = code_match.group(1)
                name = link.text.strip()
                
                price_text = tds[2].text.strip().replace(',', '')
                price = int(price_text) if price_text.isdigit() else None
                
                change_pct = None
                if len(tds) > 4:
                    try:
                        pct_text = tds[4].text.strip().replace('%', '').replace('+', '')
                        change_pct = float(pct_text) if pct_text else None
                    except:
                        pass
                
                market_cap = None
                if len(tds) > 6:
                    try:
                        cap_text = tds[6].text.strip().replace(',', '')
                        market_cap = int(cap_text) if cap_text.isdigit() else None
                    except:
                        pass
                
                stocks.append({
                    'code': code, 'name': name, 'price': price,
                    'change_pct': change_pct, 'market_cap': market_cap, 'market': market
                })
            except:
                continue
        
        time.sleep(0.2)
    
    log(f"  네이버 {market}: {len(stocks)}개 로드")
    return stocks

def get_naver_stock_detail(code):
    """네이버에서 종목 상세 정보 - 확장 버전"""
    data = {}

    # 1. 메인 페이지에서 기본 정보
    url = f"https://finance.naver.com/item/main.naver?code={code}"
    resp = fetch_naver(url)

    if resp:
        try:
            soup = BeautifulSoup(resp.text, 'lxml')

            # PER, PBR, 배당수익률
            for em in soup.select('em'):
                em_id = em.get('id', '')
                try:
                    if em_id == '_per':
                        val = float(em.text.replace(',', ''))
                        if 0 < val < 1000:
                            data['per'] = val
                    elif em_id == '_pbr':
                        val = float(em.text.replace(',', ''))
                        if 0 < val < 100:
                            data['pbr'] = val
                    elif em_id == '_dvr':
                        val = float(em.text.replace(',', ''))
                        if 0 <= val < 50:
                            data['div_yield'] = val
                    elif em_id == '_eps':
                        val = float(em.text.replace(',', ''))
                        data['eps'] = val
                    elif em_id == '_bps':
                        val = float(em.text.replace(',', ''))
                        data['bps'] = val
                except:
                    pass

            # 현재가
            try:
                price_elem = soup.select_one('p.no_today span.blind')
                if price_elem:
                    data['price'] = int(price_elem.text.replace(',', ''))
            except:
                pass

            # 시가총액
            try:
                for em in soup.select('em#_market_sum'):
                    cap_text = em.text.strip().replace(',', '').replace('조', '').replace('억', '')
                    # 시가총액은 억 단위로 저장
                    if '조' in em.text:
                        data['market_cap'] = int(float(cap_text) * 10000)
                    else:
                        data['market_cap'] = int(cap_text)
            except:
                pass

            # 외국인 비율
            try:
                for table in soup.select('table'):
                    text = table.get_text()
                    if '외국인' in text:
                        match = re.search(r'외국인[^%\d]*([\d.]+)\s*%', text)
                        if match:
                            val = float(match.group(1))
                            if 0 <= val <= 100:
                                data['foreign_ratio'] = val
                                break
            except:
                pass

            # 52주 고저
            try:
                for table in soup.select('table'):
                    for tr in table.select('tr'):
                        th = tr.select_one('th')
                        td = tr.select_one('td')
                        if th and td:
                            th_text = th.get_text().strip()
                            td_text = td.get_text().strip().replace(',', '')
                            if '52주최고' in th_text or '52주 최고' in th_text:
                                match = re.search(r'(\d+)', td_text)
                                if match:
                                    data['high_52w'] = int(match.group(1))
                            elif '52주최저' in th_text or '52주 최저' in th_text:
                                match = re.search(r'(\d+)', td_text)
                                if match:
                                    data['low_52w'] = int(match.group(1))
            except:
                pass

            # 업종(섹터)
            try:
                for a in soup.select('a[href*="upjong"]'):
                    sector_text = a.get_text().strip()
                    if sector_text and len(sector_text) > 1:
                        data['sector'] = sector_text
                        break
            except:
                pass

        except:
            pass

    # 2. 투자지표 페이지에서 ROE, ROA 등 가져오기 (sleep 제거)
    url2 = f"https://finance.naver.com/item/coinfo.naver?code={code}&target=finsum_more"
    resp2 = fetch_naver(url2)

    if resp2:
        try:
            soup2 = BeautifulSoup(resp2.text, 'lxml')

            # iframe 내용이 있을 수 있으므로 테이블에서 직접 추출
            for table in soup2.select('table'):
                rows = table.select('tr')
                for row in rows:
                    cells = row.select('td, th')
                    if len(cells) >= 2:
                        label = cells[0].get_text().strip()

                        # 가장 최근 값 (보통 마지막 td)
                        for cell in reversed(cells[1:]):
                            val_text = cell.get_text().strip().replace(',', '').replace('%', '')
                            if val_text and val_text != '-' and val_text != 'N/A':
                                try:
                                    val = float(val_text)
                                    if 'ROE' in label and 'roe' not in data:
                                        if -100 < val < 200:
                                            data['roe'] = val
                                    elif 'ROA' in label and 'roa' not in data:
                                        if -100 < val < 100:
                                            data['roa'] = val
                                    elif '영업이익률' in label and 'op_margin' not in data:
                                        if -100 < val < 100:
                                            data['op_margin'] = val
                                    elif '순이익률' in label and 'net_margin' not in data:
                                        if -100 < val < 100:
                                            data['net_margin'] = val
                                    elif '부채비율' in label and 'debt_ratio' not in data:
                                        if 0 <= val < 1000:
                                            data['debt_ratio'] = val
                                    elif '유동비율' in label and 'current_ratio' not in data:
                                        if 0 < val < 1000:
                                            data['current_ratio'] = val / 100  # 백분율 → 배수
                                    elif '매출액증가율' in label and 'revenue_growth' not in data:
                                        if -100 < val < 500:
                                            data['revenue_growth'] = val
                                    elif '영업이익증가율' in label and 'op_growth' not in data:
                                        if -200 < val < 1000:
                                            data['op_growth'] = val
                                    break
                                except:
                                    pass
        except:
            pass

    return data

# ============================================================
# FnGuide 스크래핑
# ============================================================
def get_fnguide_data(code):
    """FnGuide에서 재무 데이터 가져오기"""
    global FNGUIDE_AVAILABLE
    
    # 이미 차단된 경우 스킵
    if FNGUIDE_AVAILABLE is False:
        return {}
    
    data = {}
    url = f"https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A{code}"

    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_SHORT)
        if resp.status_code == 403:
            FNGUIDE_AVAILABLE = False
            log("  ⚠️ FnGuide 접근 차단됨")
            return data
        if resp.status_code != 200:
            return data
        
        FNGUIDE_AVAILABLE = True
        soup = BeautifulSoup(resp.text, 'lxml')

        # 시가총액, PER, PBR 등 기본 정보
        try:
            for table in soup.select('table.us_table_ty1'):
                for tr in table.select('tr'):
                    tds = tr.select('td')
                    ths = tr.select('th')
                    if len(ths) >= 1 and len(tds) >= 1:
                        for i, th in enumerate(ths):
                            label = th.get_text().strip()
                            if i < len(tds):
                                val_text = tds[i].get_text().strip().replace(',', '').replace('%', '')
                                if val_text and val_text != '-':
                                    try:
                                        val = float(val_text)
                                        if 'PER' in label and 'per' not in data:
                                            if 0 < val < 1000:
                                                data['per'] = val
                                        elif 'PBR' in label and 'pbr' not in data:
                                            if 0 < val < 100:
                                                data['pbr'] = val
                                        elif 'ROE' in label and 'roe' not in data:
                                            if -100 < val < 200:
                                                data['roe'] = val
                                        elif 'ROA' in label and 'roa' not in data:
                                            if -100 < val < 100:
                                                data['roa'] = val
                                    except:
                                        pass
        except:
            pass

        # 재무비율 테이블에서 추가 정보
        try:
            for table in soup.select('table'):
                text = table.get_text()
                if '영업이익률' in text or '부채비율' in text:
                    for tr in table.select('tr'):
                        cells = tr.select('td, th')
                        if len(cells) >= 2:
                            label = cells[0].get_text().strip()

                            # 최신 값 추출 (마지막 유효한 값)
                            for cell in reversed(cells[1:]):
                                val_text = cell.get_text().strip().replace(',', '').replace('%', '')
                                if val_text and val_text != '-' and val_text != 'N/A':
                                    try:
                                        val = float(val_text)
                                        if '영업이익률' in label and 'op_margin' not in data:
                                            data['op_margin'] = val
                                        elif '순이익률' in label and 'net_margin' not in data:
                                            data['net_margin'] = val
                                        elif '부채비율' in label and 'debt_ratio' not in data:
                                            if 0 <= val < 1000:
                                                data['debt_ratio'] = val
                                        elif '유동비율' in label and 'current_ratio' not in data:
                                            if val > 0:
                                                data['current_ratio'] = val / 100
                                        elif '매출액증가율' in label and 'revenue_growth' not in data:
                                            data['revenue_growth'] = val
                                        elif 'EPS' in label and 'eps' not in data:
                                            data['eps'] = val
                                        elif 'BPS' in label and 'bps' not in data:
                                            data['bps'] = val
                                        break
                                    except:
                                        pass
        except:
            pass

    except:
        pass

    return data

def get_naver_etf_list(max_pages=5):
    """네이버에서 ETF 리스트"""
    etfs = []
    log("  네이버 ETF 리스트 로드 시도...")
    
    for page in range(1, max_pages + 1):
        url = f"https://finance.naver.com/sise/etf.naver?page={page}"
        resp = fetch_naver(url)
        
        if not resp:
            if NAVER_AVAILABLE is False:
                return []
            continue
        
        soup = BeautifulSoup(resp.text, 'lxml')
        
        for table in soup.select('table.type_1, table.type_2'):
            for tr in table.select('tr'):
                tds = tr.select('td')
                if len(tds) < 5:
                    continue
                
                try:
                    link = tds[0].select_one('a')
                    if not link:
                        continue
                    
                    href = link.get('href', '')
                    code_match = re.search(r'code=(\d{6})', href)
                    if not code_match:
                        continue
                    
                    code = code_match.group(1)
                    name = link.text.strip()
                    
                    price = None
                    if len(tds) > 1:
                        price_text = tds[1].text.strip().replace(',', '')
                        price = int(price_text) if price_text.isdigit() else None
                    
                    etfs.append({'code': code, 'name': name, 'price': price})
                except:
                    continue
        
        time.sleep(0.1)
    
    log(f"  네이버 ETF: {len(etfs)}개 로드")
    return etfs

def get_krx_etf_data():
    """KRX에서 ETF 전종목 데이터 직접 수집 (가장 정확)"""
    etf_data = {}
    
    try:
        log("  KRX에서 ETF 전종목 로드 중...")
        
        # 1. ETF 전종목 시세
        gen_otp_url = 'http://data.krx.co.kr/comm/fileDn/GenerateOTP/generate.cmd'
        gen_otp_data = {
            'locale': 'ko_KR',
            'mktId': 'ETF',
            'trdDd': datetime.now().strftime('%Y%m%d'),
            'share': '1',
            'money': '1',
            'csvxls_isNo': 'false',
            'name': 'fileDown',
            'url': 'dbms/MDC/STAT/standard/MDCSTAT04301'
        }
        
        otp_resp = requests.post(gen_otp_url, data=gen_otp_data, headers=HEADERS, timeout=10)
        otp = otp_resp.text
        
        down_url = 'http://data.krx.co.kr/comm/fileDn/download_csv/download.cmd'
        down_resp = requests.post(
            down_url, 
            data={'code': otp}, 
            headers={**HEADERS, 'Referer': gen_otp_url},
            timeout=30
        )
        
        from io import StringIO
        csv_text = down_resp.content.decode('euc-kr', errors='ignore')
        df = pd.read_csv(StringIO(csv_text))
        
        if not df.empty and len(df) > 50:
            log(f"  KRX: {len(df)}개 ETF 로드")
            
            for _, row in df.iterrows():
                code = str(row.get('종목코드', '')).strip()
                if not code or len(code) != 6:
                    continue
                    
                etf_data[code] = {
                    'name': str(row.get('종목명', '')).strip(),
                    'price': row.get('종가', row.get('현재가', None)),
                    'nav': row.get('NAV', row.get('순자산가치', None)),
                    'volume': row.get('거래량', None),
                }
            
            return etf_data
                
    except Exception as e:
        log(f"  ⚠️ KRX 크롤링 실패: {e}")
    
    return {}

# ============================================================
# 기술적 지표 계산
# ============================================================
def calc_rsi(prices, period=14):
    try:
        delta = prices.diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi.iloc[-1]
    except:
        return None

def calc_volatility(prices, period=20):
    try:
        returns = prices.pct_change().dropna()
        if len(returns) < period:
            return None
        vol = returns.rolling(window=period).std().iloc[-1]
        return vol * np.sqrt(252) * 100
    except:
        return None

def calc_bollinger_position(prices, period=20):
    try:
        ma = prices.rolling(window=period).mean()
        std = prices.rolling(window=period).std()
        upper = ma + 2 * std
        lower = ma - 2 * std
        current = prices.iloc[-1]
        position = (current - lower.iloc[-1]) / (upper.iloc[-1] - lower.iloc[-1]) * 100
        return max(0, min(100, position))
    except:
        return None

def calc_max_drawdown(prices):
    try:
        peak = prices.expanding().max()
        drawdown = (prices - peak) / peak * 100
        return drawdown.min()
    except:
        return None

def calc_sharpe_ratio(prices, period=252, risk_free=0.02):
    try:
        returns = prices.pct_change().dropna()
        if len(returns) < period:
            return None
        mean_return = returns.mean() * 252
        std_return = returns.std() * np.sqrt(252)
        if std_return == 0:
            return None
        return (mean_return - risk_free) / std_return
    except:
        return None

# ============================================================
# 데이터 검증
# ============================================================
def validate_foreign_ratio(value):
    if value is None:
        return None
    try:
        v = float(value)
        if v < 0 or v > 100:
            return None
        if v > 70:
            return None
        return v
    except:
        return None

def is_etf_stock(name, code=None):
    if not name:
        return False
    name_upper = name.upper()
    etf_brands = ['KODEX', 'TIGER', 'KBSTAR', 'ARIRANG', 'KOSEF', 'KINDEX',
                  'HANARO', 'SOL', 'ACE', 'TIMEFOLIO', 'WOORI', 'VITA',
                  'FOCUS', 'MASTER', 'TREX', 'SMART', 'PLUS', 'RISE']
    for brand in etf_brands:
        if brand.upper() in name_upper:
            return True
    etf_keywords = ['ETF', 'ETN', '레버리지', '인버스', '선물', '지수', '액티브', '채권']
    for kw in etf_keywords:
        if kw in name or kw.upper() in name_upper:
            return True
    return False

def calc_data_quality_score(row, required_cols, data_type='stock'):
    valid_count = 0
    for col in required_cols:
        if col in row and row[col] is not None:
            valid_count += 1
    quality = (valid_count / len(required_cols)) * 100 if required_cols else 100
    row['DataQuality(%)'] = fmt(quality, 0)
    return row

# ============================================================
# 한국 주식 수집
# ============================================================
def get_korea_stocks():
    """한국 주식 데이터 수집"""
    log("\n[1/5] 한국 주식 수집 중...")
    
    all_tickers = []
    
    if PYKRX_AVAILABLE:
        try:
            log("  pykrx에서 종목 리스트 로드 중...")
            kospi_tickers = pykrx_stock.get_market_ticker_list(market="KOSPI")
            
            for ticker in kospi_tickers:
                try:
                    name = pykrx_stock.get_market_ticker_name(ticker)
                except:
                    name = ticker
                if not is_etf_stock(name, ticker):
                    all_tickers.append((ticker, name, 'KOSPI'))
            
            log(f"  pykrx: {len(all_tickers)}개 종목 로드 (KOSPI만)")
        except Exception as e:
            log(f"  ⚠️ pykrx 실패: {e}")
            all_tickers = []
    
    if not all_tickers:
        log("  네이버에서 종목 리스트 로드 시도...")
        kospi_stocks = get_naver_stock_list('KOSPI', max_pages=10 if TOP_N_KR is None else 3)
        
        for stock in kospi_stocks:
            if not is_etf_stock(stock['name'], stock['code']):
                all_tickers.append((stock['code'], stock['name'], 'KOSPI'))
    
    if not all_tickers:
        log("  ❌ 종목 리스트를 가져올 수 없음")
        return pd.DataFrame()
    
    if TOP_N_KR:
        all_tickers = all_tickers[:TOP_N_KR]
    
    log(f"  대상: {len(all_tickers)}개")
    
    results = []
    start_time = time.time()  # ★ 시간 측정 시작
    
    for i, (ticker, name, market) in enumerate(all_tickers):
        try:
            row = {'Code': ticker, 'Name': name, 'Market': market}

            # ========================================
            # 1. 네이버 (주요 소스)
            # ========================================
            naver_data = {}
            if NAVER_AVAILABLE is not False:
                naver_data = get_naver_stock_detail(ticker)
                if naver_data:
                    # 기본 정보
                    if naver_data.get('price'):
                        row['Price'] = naver_data['price']
                    if naver_data.get('market_cap'):
                        row['MarketCap(억)'] = naver_data['market_cap']
                    if naver_data.get('sector'):
                        row['Sector'] = naver_data['sector']

                    # 밸류에이션
                    if naver_data.get('per'):
                        row['PER'] = fmt(naver_data['per'])
                    if naver_data.get('pbr'):
                        row['PBR'] = fmt(naver_data['pbr'])
                    if naver_data.get('eps'):
                        row['EPS'] = fmt(naver_data['eps'], 0)
                    if naver_data.get('bps'):
                        row['BPS'] = fmt(naver_data['bps'], 0)

                    # 수익성
                    if naver_data.get('roe'):
                        row['ROE(%)'] = fmt(naver_data['roe'])
                    if naver_data.get('roa'):
                        row['ROA(%)'] = fmt(naver_data['roa'])
                    if naver_data.get('op_margin'):
                        row['OpMargin(%)'] = fmt(naver_data['op_margin'])
                    if naver_data.get('net_margin'):
                        row['NetMargin(%)'] = fmt(naver_data['net_margin'])

                    # 성장성
                    if naver_data.get('revenue_growth'):
                        row['RevenueGrowth(%)'] = fmt(naver_data['revenue_growth'])
                    if naver_data.get('op_growth'):
                        row['EarningsGrowth(%)'] = fmt(naver_data['op_growth'])

                    # 안정성
                    if naver_data.get('debt_ratio'):
                        row['DebtRatio(%)'] = fmt(naver_data['debt_ratio'])
                    if naver_data.get('current_ratio'):
                        row['CurrentRatio'] = fmt(naver_data['current_ratio'])

                    # 기타
                    if naver_data.get('foreign_ratio'):
                        row['ForeignRatio(%)'] = fmt(naver_data['foreign_ratio'])
                    if naver_data.get('div_yield'):
                        row['DivYield(%)'] = fmt(naver_data['div_yield'])
                    if naver_data.get('high_52w'):
                        row['52wHigh'] = naver_data['high_52w']
                    if naver_data.get('low_52w'):
                        row['52wLow'] = naver_data['low_52w']

                # sleep 제거 (네이버 내부에서 이미 처리)

            # ========================================
            # 2. FnGuide (결측값 폴백)
            # ========================================
            # 주요 지표가 결측인 경우에만 FnGuide 호출
            need_fnguide = (
                row.get('ROE(%)') is None or
                row.get('ROA(%)') is None or
                row.get('OpMargin(%)') is None or
                row.get('DebtRatio(%)') is None
            )

            if need_fnguide:
                fnguide_data = get_fnguide_data(ticker)
                if fnguide_data:
                    if row.get('PER') is None and fnguide_data.get('per'):
                        row['PER'] = fmt(fnguide_data['per'])
                    if row.get('PBR') is None and fnguide_data.get('pbr'):
                        row['PBR'] = fmt(fnguide_data['pbr'])
                    if row.get('ROE(%)') is None and fnguide_data.get('roe'):
                        row['ROE(%)'] = fmt(fnguide_data['roe'])
                    if row.get('ROA(%)') is None and fnguide_data.get('roa'):
                        row['ROA(%)'] = fmt(fnguide_data['roa'])
                    if row.get('OpMargin(%)') is None and fnguide_data.get('op_margin'):
                        row['OpMargin(%)'] = fmt(fnguide_data['op_margin'])
                    if row.get('NetMargin(%)') is None and fnguide_data.get('net_margin'):
                        row['NetMargin(%)'] = fmt(fnguide_data['net_margin'])
                    if row.get('DebtRatio(%)') is None and fnguide_data.get('debt_ratio'):
                        row['DebtRatio(%)'] = fmt(fnguide_data['debt_ratio'])
                    if row.get('CurrentRatio') is None and fnguide_data.get('current_ratio'):
                        row['CurrentRatio'] = fmt(fnguide_data['current_ratio'])
                    if row.get('RevenueGrowth(%)') is None and fnguide_data.get('revenue_growth'):
                        row['RevenueGrowth(%)'] = fmt(fnguide_data['revenue_growth'])
                    if row.get('EPS') is None and fnguide_data.get('eps'):
                        row['EPS'] = fmt(fnguide_data['eps'], 0)
                    if row.get('BPS') is None and fnguide_data.get('bps'):
                        row['BPS'] = fmt(fnguide_data['bps'], 0)
                # sleep 단축

            # ========================================
            # 3. pykrx (결측값 폴백 - 가격/거래량/시총)
            # ========================================
            if PYKRX_AVAILABLE:
                try:
                    today_str = datetime.now().strftime("%Y%m%d")
                    ohlcv = pykrx_stock.get_market_ohlcv_by_date(
                        (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
                        today_str, ticker
                    )
                    if not ohlcv.empty:
                        if row.get('Price') is None:
                            row['Price'] = fmt(ohlcv['종가'].iloc[-1], 0)
                        if row.get('Volume') is None:
                            row['Volume'] = int(ohlcv['거래량'].iloc[-1]) if ohlcv['거래량'].iloc[-1] else None
                except:
                    pass

            # ========================================
            # 4. yfinance (최후의 수단 - 기술적 지표)
            # ========================================
            hist = pd.DataFrame()
            info = {}

            # yfinance는 기술적 지표(RSI, 변동성 등)와 여전히 결측인 필드에만 사용
            # KOSPI만 사용하므로 .KS 우선
            ticker_variants = [f"{ticker}.KS"]

            t, hist, info = try_multiple_tickers(ticker_variants, max_retries=1)  # 재시도 최소화

            # 여전히 결측인 기본 정보만 폴백
            if not row.get('Price'):
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))
            if not row.get('Sector'):
                row['Sector'] = safe_get(info, 'sector')
            if row.get('MarketCap(억)') is None:
                cap = safe_get(info, 'marketCap')
                if cap:
                    row['MarketCap(억)'] = fmt(cap / 1e8, 0)
            
            # 밸류에이션
            if row.get('PER') is None:
                row['PER'] = fmt(safe_get(info, 'trailingPE'))
            if row.get('ForwardPE') is None:  # ← PWA 호환: ForwardPE
                row['ForwardPE'] = fmt(safe_get(info, 'forwardPE'))
            if row.get('PBR') is None:
                row['PBR'] = fmt(safe_get(info, 'priceToBook'))
            
            # 수익성
            if row.get('ROE(%)') is None and safe_get(info, 'returnOnEquity'):
                row['ROE(%)'] = fmt(safe_get(info, 'returnOnEquity') * 100)
            if row.get('ROA(%)') is None and safe_get(info, 'returnOnAssets'):
                row['ROA(%)'] = fmt(safe_get(info, 'returnOnAssets') * 100)
            if row.get('OpMargin(%)') is None and safe_get(info, 'operatingMargins'):
                row['OpMargin(%)'] = fmt(safe_get(info, 'operatingMargins') * 100)
            if row.get('GrossMargin(%)') is None and safe_get(info, 'grossMargins'):
                row['GrossMargin(%)'] = fmt(safe_get(info, 'grossMargins') * 100)
            
            # 성장성 (결측인 경우만)
            if row.get('RevenueGrowth(%)') is None and safe_get(info, 'revenueGrowth'):
                row['RevenueGrowth(%)'] = fmt(safe_get(info, 'revenueGrowth') * 100)
            if row.get('EarningsGrowth(%)') is None and safe_get(info, 'earningsGrowth'):
                row['EarningsGrowth(%)'] = fmt(safe_get(info, 'earningsGrowth') * 100)

            # 안정성 (결측인 경우만) - PWA 호환: DebtRatio(%)
            if row.get('CurrentRatio') is None:
                row['CurrentRatio'] = fmt(safe_get(info, 'currentRatio'))
            if row.get('DebtRatio(%)') is None:
                row['DebtRatio(%)'] = fmt(safe_get(info, 'debtToEquity'))
            
            # 배당
            if row.get('DivYield(%)') is None and safe_get(info, 'dividendYield'):
                row['DivYield(%)'] = fmt(safe_get(info, 'dividendYield') * 100)
            
            # 기술적 지표
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                
                # 수익률
                if len(close) >= 2:
                    row['Return1D(%)'] = fmt((close.iloc[-1] / close.iloc[-2] - 1) * 100)
                if len(close) >= 6:
                    row['Return1W(%)'] = fmt((close.iloc[-1] / close.iloc[-6] - 1) * 100)
                if len(close) >= 22:
                    row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
                if len(close) >= 66:
                    row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
                    row['Return60D(%)'] = row['Return3M(%)']  # PWA 호환
                if len(close) >= 132:
                    row['Return6M(%)'] = fmt((close.iloc[-1] / close.iloc[-132] - 1) * 100)
                    row['Return120D(%)'] = row['Return6M(%)']  # PWA 호환
                if len(close) >= 252:
                    row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-252] - 1) * 100)
                    row['Return250D(%)'] = row['Return1Y(%)']  # ← PWA 호환: 추가
                
                # 이동평균
                row['MA20'] = fmt(close.rolling(20).mean().iloc[-1])
                row['MA60'] = fmt(close.rolling(60).mean().iloc[-1])
                row['MA120'] = fmt(close.rolling(120).mean().iloc[-1])
                
                price = close.iloc[-1]
                if row['MA20']: row['vs_MA20(%)'] = fmt((price / row['MA20'] - 1) * 100)
                if row['MA60']: row['vs_MA60(%)'] = fmt((price / row['MA60'] - 1) * 100)
                if row['MA120']: row['vs_MA120(%)'] = fmt((price / row['MA120'] - 1) * 100)
                
                # RSI
                row['RSI14'] = fmt(calc_rsi(close, 14))
                row['BB_Position'] = fmt(calc_bollinger_position(close, 20))
                
                # 52주
                if len(close) >= 252:
                    year_data = close.tail(252)
                    if not row.get('52wHigh'):
                        row['52wHigh'] = fmt(year_data.max())
                    if not row.get('52wLow'):
                        row['52wLow'] = fmt(year_data.min())
                    
                    current_price = row.get('Price') or close.iloc[-1]
                    if current_price and row.get('52wHigh'):
                        row['From52wHigh(%)'] = fmt((float(current_price) / float(row['52wHigh']) - 1) * 100)
                    if current_price and row.get('52wLow'):
                        row['From52wLow(%)'] = fmt((float(current_price) / float(row['52wLow']) - 1) * 100)
                
                # 변동성
                row['Volatility20D'] = fmt(calc_volatility(close, 20))
                row['Volatility60D'] = fmt(calc_volatility(close, 60))
                
                # MaxDrawdown
                row['MaxDrawdown(%)'] = fmt(calc_max_drawdown(close))
                
                # SharpeRatio - PWA 호환
                row['SharpeRatio'] = fmt(calc_sharpe_ratio(close))  # ← PWA 호환: Sharpe1Y → SharpeRatio

            # 최종 폴백: 여전히 결측인 필드들을 info에서 다시 시도
            kr_field_mapping = {
                'PER': (('trailingPE', 'forwardPE'), 1, 2),
                'PBR': (('priceToBook',), 1, 2),
                'ROE(%)': (('returnOnEquity',), 100, 2),
                'ROA(%)': (('returnOnAssets',), 100, 2),
                'DivYield(%)': (('dividendYield',), 100, 2),
                'Beta': (('beta',), 1, 2),
                'MarketCap(억)': (('marketCap',), 1e-8, 0),
            }
            row = fill_missing_from_info(row, info, kr_field_mapping)

            # 52주 고저 폴백
            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            row = calc_data_quality_score(row, REQUIRED_COLS_KR_STOCK, 'kr_stock')
            results.append(row)
            
        except Exception as e:
            results.append({'Code': ticker, 'Name': name, 'Market': market, 'Remark': str(e)[:30]})
        
        # ★ 진행 상황 출력 (10개마다 + 예상 시간)
        if (i + 1) % 10 == 0 or i == 0:
            elapsed = time.time() - start_time
            per_stock = elapsed / (i + 1) if i > 0 else 0
            remaining = per_stock * (len(all_tickers) - i - 1)
            log(f"  진행: {i+1}/{len(all_tickers)} ({(i+1)/len(all_tickers)*100:.0f}%) - 남은시간: {remaining/60:.1f}분")
        
        time.sleep(0.02)  # 0.05 → 0.02
    
    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# 미국 주식 수집
# ============================================================
def get_us_stocks():
    """미국 S&P 500 주식 데이터"""
    log("\n[2/5] 미국 주식 수집 중...")
    
    sp500_top = [
        'AAPL', 'MSFT', 'AMZN', 'NVDA', 'GOOGL', 'META', 'TSLA', 'BRK-B', 'UNH', 'JNJ',
        'XOM', 'JPM', 'V', 'PG', 'MA', 'HD', 'CVX', 'MRK', 'ABBV', 'LLY',
        'PEP', 'KO', 'COST', 'AVGO', 'TMO', 'MCD', 'WMT', 'CSCO', 'ACN', 'ABT',
        'DHR', 'NEE', 'VZ', 'ADBE', 'CRM', 'NKE', 'PM', 'TXN', 'WFC', 'BMY',
        'RTX', 'CMCSA', 'ORCL', 'UPS', 'HON', 'QCOM', 'COP', 'T', 'LOW', 'MS',
        'INTC', 'UNP', 'ELV', 'BA', 'SPGI', 'CAT', 'IBM', 'GS', 'PLD', 'INTU',
        'DE', 'AMD', 'SBUX', 'AXP', 'AMAT', 'MDLZ', 'GE', 'BLK', 'GILD', 'ADI',
        'LMT', 'ISRG', 'TJX', 'SYK', 'CVS', 'BKNG', 'ADP', 'MMC', 'VRTX', 'REGN',
        'PGR', 'CB', 'NOW', 'CI', 'SCHW', 'ZTS', 'MO', 'TMUS', 'SO', 'DUK',
        'EOG', 'BDX', 'C', 'PNC', 'CL', 'TGT', 'ITW', 'SLB', 'AMT', 'USB',
    ]
    
    tickers = sp500_top[:TOP_N_US] if TOP_N_US else sp500_top
    results = []
    
    log(f"  대상: {len(tickers)}개")
    
    for i, ticker in enumerate(tickers):
        try:
            # 재시도 로직과 함께 데이터 가져오기
            t, hist, info = fetch_yf_with_retry(ticker, max_retries=3, delay=0.5)

            if hist.empty and not info:
                continue

            row = {'Ticker': ticker}
            
            row['Name'] = safe_get(info, 'shortName', 'longName')
            row['Sector'] = safe_get(info, 'sector')
            row['Industry'] = safe_get(info, 'industry')
            row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))
            
            cap = safe_get(info, 'marketCap')
            if cap:
                row['MarketCap(B)'] = fmt(cap / 1e9, 1)
            
            # 밸류에이션 - PWA 호환
            row['PER'] = fmt(safe_get(info, 'trailingPE'))
            row['ForwardPE'] = fmt(safe_get(info, 'forwardPE'))  # ← PWA 호환: ForwardPE
            row['PBR'] = fmt(safe_get(info, 'priceToBook'))
            row['PSR'] = fmt(safe_get(info, 'priceToSalesTrailing12Months'))
            row['PEG'] = fmt(safe_get(info, 'pegRatio'))
            row['EV/EBITDA'] = fmt(safe_get(info, 'enterpriseToEbitda'))
            
            # 수익성
            row['ROE(%)'] = fmt(safe_get(info, 'returnOnEquity', default=0) * 100) if safe_get(info, 'returnOnEquity') else None
            row['ROA(%)'] = fmt(safe_get(info, 'returnOnAssets', default=0) * 100) if safe_get(info, 'returnOnAssets') else None
            row['GrossMargin(%)'] = fmt(safe_get(info, 'grossMargins', default=0) * 100) if safe_get(info, 'grossMargins') else None
            row['OpMargin(%)'] = fmt(safe_get(info, 'operatingMargins', default=0) * 100) if safe_get(info, 'operatingMargins') else None
            row['NetMargin(%)'] = fmt(safe_get(info, 'profitMargins', default=0) * 100) if safe_get(info, 'profitMargins') else None
            
            # FCF Yield
            fcf = safe_get(info, 'freeCashflow')
            cap = safe_get(info, 'marketCap')
            if fcf and cap and cap > 0:
                row['FCFYield(%)'] = fmt(fcf / cap * 100)
            
            # 성장성
            row['RevenueGrowth(%)'] = fmt(safe_get(info, 'revenueGrowth', default=0) * 100) if safe_get(info, 'revenueGrowth') else None
            row['EarningsGrowth(%)'] = fmt(safe_get(info, 'earningsGrowth', default=0) * 100) if safe_get(info, 'earningsGrowth') else None
            
            # 안정성 - PWA 호환
            row['CurrentRatio'] = fmt(safe_get(info, 'currentRatio'))
            row['QuickRatio'] = fmt(safe_get(info, 'quickRatio'))
            row['DebtRatio(%)'] = fmt(safe_get(info, 'debtToEquity'))  # ← PWA 호환
            row['Debt/Equity'] = fmt(safe_get(info, 'debtToEquity'))   # ← PWA 호환: 추가
            
            # 배당
            row['DivYield(%)'] = fmt(safe_get(info, 'dividendYield', default=0) * 100) if safe_get(info, 'dividendYield') else None
            row['PayoutRatio(%)'] = fmt(safe_get(info, 'payoutRatio', default=0) * 100) if safe_get(info, 'payoutRatio') else None
            
            # 베타
            row['Beta'] = fmt(safe_get(info, 'beta'))
            
            # 기관 보유 비율 - PWA 호환: 추가
            inst = safe_get(info, 'heldPercentInstitutions')
            if inst:
                row['InstOwn(%)'] = fmt(inst * 100)  # ← PWA 호환: 추가
            
            # 기술적 지표
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                
                # 수익률 - PWA 호환
                if len(close) >= 2:
                    row['Return1D(%)'] = fmt((close.iloc[-1] / close.iloc[-2] - 1) * 100)
                if len(close) >= 6:
                    row['Return1W(%)'] = fmt((close.iloc[-1] / close.iloc[-6] - 1) * 100)
                if len(close) >= 22:
                    row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
                if len(close) >= 66:
                    row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
                    row['Return60D(%)'] = row['Return3M(%)']  # PWA 호환
                if len(close) >= 132:
                    row['Return6M(%)'] = fmt((close.iloc[-1] / close.iloc[-132] - 1) * 100)
                    row['Return120D(%)'] = row['Return6M(%)']  # PWA 호환
                if len(close) >= 252:
                    row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-252] - 1) * 100)
                    row['Return250D(%)'] = row['Return1Y(%)']  # ← PWA 호환: 추가
                
                # YTD
                try:
                    ytd_start = close[close.index >= f"{datetime.now().year}-01-01"]
                    if len(ytd_start) > 1:
                        row['ReturnYTD(%)'] = fmt((close.iloc[-1] / ytd_start.iloc[0] - 1) * 100)
                except:
                    pass
                
                # 이동평균
                row['MA20'] = fmt(close.rolling(20).mean().iloc[-1])
                row['MA50'] = fmt(close.rolling(50).mean().iloc[-1])
                row['MA200'] = fmt(close.rolling(200).mean().iloc[-1])
                
                price = close.iloc[-1]
                if row['MA20']: row['vs_MA20(%)'] = fmt((price / row['MA20'] - 1) * 100)
                if row['MA50']: row['vs_MA50(%)'] = fmt((price / row['MA50'] - 1) * 100)
                if row['MA200']: row['vs_MA200(%)'] = fmt((price / row['MA200'] - 1) * 100)
                
                # RSI
                row['RSI14'] = fmt(calc_rsi(close, 14))
                row['BB_Position'] = fmt(calc_bollinger_position(close, 20))
                
                # 52주 고저
                if len(close) >= 252:
                    year_data = close.tail(252)
                    row['52wHigh'] = fmt(year_data.max())
                    row['52wLow'] = fmt(year_data.min())
                    row['From52wHigh(%)'] = fmt((price / year_data.max() - 1) * 100)
                    row['From52wLow(%)'] = fmt((price / year_data.min() - 1) * 100)
                
                # 변동성
                row['Volatility20D'] = fmt(calc_volatility(close, 20))
                row['Volatility60D'] = fmt(calc_volatility(close, 60))
                
                # MaxDrawdown
                row['MaxDrawdown(%)'] = fmt(calc_max_drawdown(close))
                
                # SharpeRatio - PWA 호환
                row['SharpeRatio'] = fmt(calc_sharpe_ratio(close))  # ← PWA 호환

            # 최종 폴백: 여전히 결측인 필드들을 info에서 다시 시도
            us_field_mapping = {
                'PER': (('trailingPE', 'forwardPE'), 1, 2),
                'ForwardPE': (('forwardPE',), 1, 2),
                'PBR': (('priceToBook',), 1, 2),
                'ROE(%)': (('returnOnEquity',), 100, 2),
                'ROA(%)': (('returnOnAssets',), 100, 2),
                'DivYield(%)': (('dividendYield',), 100, 2),
                'Beta': (('beta',), 1, 2),
                'GrossMargin(%)': (('grossMargins',), 100, 2),
                'OpMargin(%)': (('operatingMargins',), 100, 2),
                'NetMargin(%)': (('profitMargins',), 100, 2),
            }
            row = fill_missing_from_info(row, info, us_field_mapping)

            # 52주 고저 폴백
            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            # Price 폴백
            if row.get('Price') is None:
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice', 'previousClose', 'open'))

            row = calc_data_quality_score(row, REQUIRED_COLS_US_STOCK, 'us_stock')
            results.append(row)

        except Exception as e:
            results.append({'Ticker': ticker, 'Remark': str(e)[:30]})

        if (i + 1) % 20 == 0:
            log(f"  진행: {i+1}/{len(tickers)}")

        time.sleep(0.15)

    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# ETF 공통 함수
# ============================================================
def get_etf_data(tickers, region=""):
    """ETF 데이터 수집 공통 함수"""
    results = []

    for i, ticker in enumerate(tickers):
        try:
            # 재시도 로직과 함께 데이터 가져오기
            t, hist, info = fetch_yf_with_retry(ticker, max_retries=3, delay=0.5)

            if hist.empty and not info:
                continue
            
            row = {'Ticker': ticker, 'Region': region}
            
            row['Name'] = safe_get(info, 'shortName', 'longName')
            row['Category'] = safe_get(info, 'category')
            row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))
            
            row['ExpenseRatio(%)'] = fmt(safe_get(info, 'expenseRatio', default=0) * 100) if safe_get(info, 'expenseRatio') else None
            row['TotalAssets(B)'] = fmt(safe_get(info, 'totalAssets', default=0) / 1e9, 2) if safe_get(info, 'totalAssets') else None
            
            row['DivYield(%)'] = fmt(safe_get(info, 'yield', default=0) * 100) if safe_get(info, 'yield') else None
            
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                
                # 수익률 - PWA 호환 별칭 추가
                if len(close) >= 2:
                    row['Return1D(%)'] = fmt((close.iloc[-1] / close.iloc[-2] - 1) * 100)
                if len(close) >= 6:
                    row['Return1W(%)'] = fmt((close.iloc[-1] / close.iloc[-6] - 1) * 100)
                if len(close) >= 22:
                    row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
                if len(close) >= 66:
                    row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
                    row['Return60D(%)'] = row['Return3M(%)']  # ← PWA 호환: 추가
                if len(close) >= 132:
                    row['Return6M(%)'] = fmt((close.iloc[-1] / close.iloc[-132] - 1) * 100)
                    row['Return120D(%)'] = row['Return6M(%)']  # ← PWA 호환: 추가
                if len(close) >= 252:
                    row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-252] - 1) * 100)
                    row['Return250D(%)'] = row['Return1Y(%)']  # ← PWA 호환: 추가
                
                row['MA20'] = fmt(close.rolling(20).mean().iloc[-1])
                row['MA50'] = fmt(close.rolling(50).mean().iloc[-1])
                
                row['RSI14'] = fmt(calc_rsi(close, 14))
                
                row['Volatility20D'] = fmt(calc_volatility(close, 20))
                row['Volatility60D'] = fmt(calc_volatility(close, 60))
                
                if len(close) >= 252:
                    year_data = close.tail(252)
                    row['52wHigh'] = fmt(year_data.max())
                    row['52wLow'] = fmt(year_data.min())
                    row['From52wHigh(%)'] = fmt((close.iloc[-1] / year_data.max() - 1) * 100)
                    row['From52wLow(%)'] = fmt((close.iloc[-1] / year_data.min() - 1) * 100)
                
                row['MaxDrawdown(%)'] = fmt(calc_max_drawdown(close))
                row['SharpeRatio'] = fmt(calc_sharpe_ratio(close))  # ← PWA 호환

            # 최종 폴백: 여전히 결측인 필드들을 info에서 다시 시도
            etf_field_mapping = {
                'ExpenseRatio(%)': (('expenseRatio',), 100, 3),
                'DivYield(%)': (('yield', 'dividendYield'), 100, 2),
                'TotalAssets(B)': (('totalAssets',), 1e-9, 2),
            }
            row = fill_missing_from_info(row, info, etf_field_mapping)

            # 52주 고저 폴백
            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            # Price 폴백
            if row.get('Price') is None:
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice', 'previousClose', 'navPrice'))

            row = calc_data_quality_score(row, REQUIRED_COLS_ETF, 'etf')
            results.append(row)

        except Exception as e:
            results.append({'Ticker': ticker, 'Remark': str(e)[:30]})

        time.sleep(0.15)

    return pd.DataFrame(results)

# ============================================================
# 한국 ETF
# ============================================================
def get_korea_etfs():
    """한국 ETF 데이터"""
    log("\n[3/5] 한국 ETF 수집 중...")
    
    # ========================================
    # 1. KRX에서 ETF 전종목 데이터 (가장 정확)
    # ========================================
    krx_data = get_krx_etf_data()
    
    if krx_data:
        kr_etf_list = list(krx_data.keys())
        log(f"  KRX: {len(kr_etf_list)}개 ETF")
    else:
        # ========================================
        # 2. pykrx 대안
        # ========================================
        kr_etf_list = []
        if PYKRX_AVAILABLE:
            try:
                log("  pykrx에서 ETF 리스트 로드 중...")
                kr_etf_list = pykrx_stock.get_etf_ticker_list()
                log(f"  pykrx: {len(kr_etf_list)}개 ETF")
            except Exception as e:
                log(f"  ⚠️ pykrx ETF 실패: {e}")
        
        # ========================================
        # 3. 네이버 대안
        # ========================================
        if not kr_etf_list and NAVER_AVAILABLE is not False:
            naver_etfs = get_naver_etf_list(max_pages=10)
            kr_etf_list = [etf['code'] for etf in naver_etfs]
    
    # ========================================
    # 4. 하드코딩 폴백 (모두 실패 시)
    # ========================================
    if not kr_etf_list and not krx_data:
        log("  ⚠️ ETF 리스트 로드 실패, 주요 ETF만 사용")
        kr_etf_list = [
            '069500', '114800', '122630', '229200', '252670',
            '305720', '091160', '133690', '143850', '192090',
            '261240', '360750', '371450', '379800', '381170',
            '395160', '461460',
        ]
    
    # ★ ETF 개수 제한 적용
    if TOP_N_KR_ETF and len(kr_etf_list) > TOP_N_KR_ETF:
        kr_etf_list = kr_etf_list[:TOP_N_KR_ETF]
    
    log(f"  대상: {len(kr_etf_list)}개")
    
    results = []
    start_time = time.time()
    
    for i, code in enumerate(kr_etf_list):
        try:
            row = {'Code': code, 'Region': 'KR'}

            # ========================================
            # 1. KRX 데이터 먼저 적용 (있으면)
            # ========================================
            krx_info = krx_data.get(code, {}) if krx_data else {}
            if krx_info:
                row['Name'] = krx_info.get('name', '')
                if krx_info.get('price'):
                    row['Price'] = fmt(krx_info['price'], 0)
                if krx_info.get('nav'):
                    row['NAV'] = fmt(krx_info['nav'], 0)
                if krx_info.get('volume'):
                    row['Volume'] = int(krx_info['volume'])

            # ========================================
            # 2. pykrx 보충 (KRX 없으면)
            # ========================================
            if PYKRX_AVAILABLE and not row.get('Name'):
                try:
                    row['Name'] = pykrx_stock.get_etf_ticker_name(code)
                    today_str = datetime.now().strftime("%Y%m%d")
                    ohlcv = pykrx_stock.get_etf_ohlcv_by_date(
                        (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
                        today_str, code
                    )
                    if not ohlcv.empty:
                        if not row.get('Price'):
                            row['Price'] = fmt(ohlcv['종가'].iloc[-1], 0)
                        if not row.get('Volume'):
                            row['Volume'] = int(ohlcv['거래량'].iloc[-1]) if ohlcv['거래량'].iloc[-1] else None
                except:
                    pass

            # ========================================
            # 3. yfinance (기술적 지표)
            # ========================================
            ticker_variants = [f"{code}.KS", f"{code}.KQ"]
            t, hist, info = try_multiple_tickers(ticker_variants, max_retries=1)

            if not row.get('Name'):
                row['Name'] = safe_get(info, 'shortName', 'longName') or code
            if not row.get('Price'):
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))

            row['Category'] = safe_get(info, 'category')
            row['ExpenseRatio(%)'] = fmt(safe_get(info, 'expenseRatio', default=0) * 100) if safe_get(info, 'expenseRatio') else None
            
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                
                # 수익률 - PWA 호환
                if len(close) >= 22:
                    row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
                if len(close) >= 66:
                    row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
                    row['Return60D(%)'] = row['Return3M(%)']  # ← PWA 호환
                if len(close) >= 132:
                    row['Return6M(%)'] = fmt((close.iloc[-1] / close.iloc[-132] - 1) * 100)
                    row['Return120D(%)'] = row['Return6M(%)']  # ← PWA 호환
                if len(close) >= 252:
                    row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-252] - 1) * 100)
                    row['Return250D(%)'] = row['Return1Y(%)']  # ← PWA 호환
                
                row['RSI14'] = fmt(calc_rsi(close, 14))
                row['Volatility20D'] = fmt(calc_volatility(close, 20))
                row['MaxDrawdown(%)'] = fmt(calc_max_drawdown(close))
                row['SharpeRatio'] = fmt(calc_sharpe_ratio(close))  # ← PWA 호환
            else:
                # hist가 비어있어도 info가 있으면 기본 정보라도 포함
                if not info:
                    continue

            # 최종 폴백: 여전히 결측인 필드들을 info에서 다시 시도
            kr_etf_field_mapping = {
                'ExpenseRatio(%)': (('expenseRatio',), 100, 3),
                'DivYield(%)': (('yield', 'dividendYield'), 100, 2),
            }
            row = fill_missing_from_info(row, info, kr_etf_field_mapping)

            # 52주 고저 폴백
            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            row = calc_data_quality_score(row, REQUIRED_COLS_ETF, 'etf')
            results.append(row)

        except Exception as e:
            pass
        
        # 진행 상황 (20개마다)
        if (i + 1) % 20 == 0 or i == 0:
            elapsed = time.time() - start_time
            per_etf = elapsed / (i + 1) if i > 0 else 0
            remaining = per_etf * (len(kr_etf_list) - i - 1)
            log(f"  진행: {i+1}/{len(kr_etf_list)} - 남은시간: {remaining/60:.1f}분")
        
        time.sleep(0.05)
    
    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# 미국 ETF
# ============================================================
def get_us_etfs():
    """미국 ETF 데이터"""
    log("\n[4/5] 미국 ETF 수집 중...")
    
    tickers = [
        'SPY', 'IVV', 'VOO', 'VTI', 'QQQ', 'DIA', 'IWM', 'IWF', 'IWD', 'VUG',
        'VTV', 'IJH', 'IJR', 'VB', 'VO', 'RSP', 'SPLG', 'SCHX', 'SCHB', 'MGK',
        'XLK', 'XLF', 'XLV', 'XLE', 'XLI', 'XLY', 'XLP', 'XLU', 'XLB', 'XLRE',
        'VGT', 'VFH', 'VHT', 'VDE', 'VIS', 'VCR', 'VDC', 'VPU', 'VAW', 'VNQ',
        'ARKK', 'ARKW', 'ARKF', 'ARKG', 'SOXX', 'SMH', 'XBI', 'IBB', 'HACK', 'BOTZ',
        'LIT', 'TAN', 'ICLN', 'PBW', 'QCLN', 'REMX', 'COPX', 'URA',
        'VYM', 'SCHD', 'DVY', 'HDV', 'SPHD', 'SPYD', 'VIG', 'DGRO', 'NOBL',
        'BND', 'AGG', 'TLT', 'IEF', 'SHY', 'LQD', 'HYG', 'JNK', 'TIP',
        'GLD', 'IAU', 'SLV', 'USO', 'UNG', 'DBC', 'PDBC',
        'TQQQ', 'SQQQ', 'UPRO', 'SPXU', 'SOXL', 'SOXS',
        'VEA', 'VWO', 'EFA', 'EEM', 'IEFA', 'IEMG', 'VXUS', 'ACWI'
    ]
    
    # ★ ETF 개수 제한 적용
    if TOP_N_US_ETF and len(tickers) > TOP_N_US_ETF:
        tickers = tickers[:TOP_N_US_ETF]
    
    log(f"  대상: {len(tickers)}개")
    df = get_etf_data(tickers, "US")
    log(f"  ✅ 완료: {len(df)}개")
    return df

# ============================================================
# 시장 지표
# ============================================================
def get_fear_greed_index():
    """CNN Fear & Greed Index 스크래핑"""
    try:
        url = "https://production.dataviz.cnn.io/index/fearandgreed/graphdata"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
        }
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if 'fear_and_greed' in data:
                return {
                    'value': fmt(data['fear_and_greed'].get('score'), 1),
                    'rating': data['fear_and_greed'].get('rating', ''),
                    'previous_close': fmt(data['fear_and_greed'].get('previous_close'), 1),
                }
    except:
        pass
    return None

def get_cape_ratio():
    """S&P 500 Shiller CAPE Ratio from multpl.com"""
    try:
        url = "https://www.multpl.com/shiller-pe"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            # Current value
            current = soup.select_one('#current')
            if current:
                val_text = current.get_text().strip().replace(',', '')
                match = re.search(r'([\d.]+)', val_text)
                if match:
                    return fmt(float(match.group(1)), 2)
    except:
        pass
    return None

def get_sp500_forward_pe():
    """S&P 500 Forward PE from SPY ETF"""
    try:
        spy = yf.Ticker('SPY')
        info = spy.info
        # SPY doesn't have forward PE directly, estimate from holdings
        # Use IVV or VOO as alternatives
        for ticker in ['SPY', 'IVV', 'VOO']:
            t = yf.Ticker(ticker)
            info = t.info
            pe = safe_get(info, 'trailingPE')
            if pe and 10 < pe < 50:
                return fmt(pe, 2)
    except:
        pass
    return None

def get_ism_pmi():
    """ISM 제조업 PMI from tradingeconomics"""
    try:
        url = "https://tradingeconomics.com/united-states/business-confidence"
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_LONG)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            # ISM Manufacturing PMI 값 찾기
            for elem in soup.select('#aspnetForm td, #aspnetForm span'):
                text = elem.get_text().strip()
                # 숫자값 추출 (40-60 범위)
                try:
                    val = float(text)
                    if 40 <= val <= 65:  # PMI 범위
                        return fmt(val, 1)
                except:
                    pass
    except:
        pass
    
    # 대안: investing.com
    try:
        url = "https://www.investing.com/economic-calendar/ism-manufacturing-pmi-173"
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_LONG)
        if resp.status_code == 200:
            match = re.search(r'Actual.*?(\d+\.?\d*)', resp.text)
            if match:
                val = float(match.group(1))
                if 40 <= val <= 65:
                    return fmt(val, 1)
    except:
        pass
    
    return None

def get_korea_market_indicators():
    """한국 시장 지표 (pykrx + 네이버 대안)"""
    indicators = {}

    # ========================================
    # 1. pykrx 시도
    # ========================================
    if PYKRX_AVAILABLE:
        try:
            today_str = datetime.now().strftime("%Y%m%d")

            # 코스피 PER, PBR
            try:
                from_date = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d")
                kospi_fund = pykrx_stock.get_index_fundamental(from_date, today_str, "1001")
                if not kospi_fund.empty:
                    latest = kospi_fund.iloc[-1]
                    if 'PER' in kospi_fund.columns:
                        indicators['kospi_per'] = fmt(latest['PER'], 2)
                    if 'PBR' in kospi_fund.columns:
                        indicators['kospi_pbr'] = fmt(latest['PBR'], 2)
                    if 'DIV' in kospi_fund.columns:
                        indicators['kospi_div'] = fmt(latest['DIV'], 2)
            except Exception as e:
                log(f"  ⚠️ pykrx 펀더멘털 실패: {e}")

            # 외국인 순매수 (5일 + 20일)
            try:
                from_date = (datetime.now() - timedelta(days=30)).strftime("%Y%m%d")
                investor = pykrx_stock.get_market_trading_value_by_date(from_date, today_str, "KOSPI")
                if not investor.empty and '외국인' in investor.columns:
                    # 5일 순매수
                    recent_5 = investor['외국인'].tail(5).sum()
                    indicators['foreign_net_buy'] = fmt(recent_5 / 1e8, 0)
                    # 20일 누적
                    recent_20 = investor['외국인'].tail(20).sum()
                    indicators['foreign_net_buy_20d'] = fmt(recent_20 / 1e8, 0)
                    
                if not investor.empty:
                    if '개인' in investor.columns:
                        indicators['individual_net_buy'] = fmt(investor['개인'].tail(5).sum() / 1e8, 0)
                    if '기관합계' in investor.columns:
                        indicators['institution_net_buy'] = fmt(investor['기관합계'].tail(5).sum() / 1e8, 0)
            except Exception as e:
                log(f"  ⚠️ pykrx 투자자 실패: {e}")

        except Exception as e:
            log(f"  ⚠️ pykrx 전체 실패: {e}")

    # ========================================
    # 2. 네이버 대안 (pykrx 실패 시)
    # ========================================
    if not indicators.get('kospi_per') and NAVER_AVAILABLE is not False:
        try:
            # 네이버 시장 지표 페이지에서 코스피 PER
            url = "https://finance.naver.com/sise/sise_index.naver?code=KOSPI"
            resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_SHORT)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'lxml')
                # PER 찾기
                for td in soup.select('td'):
                    text = td.get_text().strip()
                    if 'PER' in text or 'per' in text.lower():
                        # 다음 td나 같은 행에서 값 추출
                        parent = td.find_parent('tr')
                        if parent:
                            tds = parent.select('td')
                            for t in tds:
                                val_text = t.get_text().strip().replace(',', '')
                                try:
                                    val = float(val_text)
                                    if 5 < val < 50:  # PER 범위
                                        indicators['kospi_per'] = fmt(val, 2)
                                        break
                                except:
                                    pass
                log("  ✅ 네이버에서 코스피 지표 로드")
        except Exception as e:
            log(f"  ⚠️ 네이버 코스피 지표 실패: {e}")

    return indicators

def get_market_indicators():
    """글로벌 시장 지표 - 확장 버전"""
    log("\n[5/5] 시장 지표 수집 중...")

    # ========================================
    # 1. Yahoo Finance 기본 지표
    # ========================================
    indicators = {
        # 지수
        '^GSPC': ('S&P 500', '지수'),
        '^DJI': ('다우존스', '지수'),
        '^IXIC': ('나스닥 종합', '지수'),
        '^VIX': ('VIX 변동성', '지수'),
        '^KS11': ('코스피', '지수'),
        '^KQ11': ('코스닥', '지수'),
        '^N225': ('니케이 225', '지수'),
        '^HSI': ('항셍', '지수'),
        '^STOXX50E': ('유로스톡스50', '지수'),
        '^FTSE': ('FTSE 100', '지수'),

        # 환율
        'USDKRW=X': ('USD/KRW', '환율'),
        'EURUSD=X': ('EUR/USD', '환율'),
        'USDJPY=X': ('USD/JPY', '환율'),
        'DX-Y.NYB': ('달러 인덱스', '환율'),
        'USDCNY=X': ('USD/CNY', '환율'),

        # 원자재
        'GC=F': ('금 선물', '원자재'),
        'SI=F': ('은 선물', '원자재'),
        'CL=F': ('WTI 원유', '원자재'),
        'BZ=F': ('브렌트유', '원자재'),
        'NG=F': ('천연가스', '원자재'),
        'HG=F': ('구리 선물', '원자재'),

        # 채권 - Tier 1 추가
        '^IRX': ('미국채 3개월', '채권'),
        '^FVX': ('미국채 5년', '채권'),
        '^TNX': ('미국채 10년', '채권'),
        '^TYX': ('미국채 30년', '채권'),

        # 암호화폐
        'BTC-USD': ('비트코인', '암호화폐'),
        'ETH-USD': ('이더리움', '암호화폐'),

        # 신용/스프레드 관련 ETF
        'HYG': ('하이일드 채권 ETF', '신용'),
        'LQD': ('투자등급 채권 ETF', '신용'),
        'TLT': ('장기국채 ETF', '채권'),
        'SHY': ('단기국채 ETF', '채권'),

        # 옵션 심리 - Tier 2
        '^CPCE': ('Put/Call Ratio', '심리'),
    }

    results = []

    log("  Yahoo Finance 지표 수집 중...")
    for ticker, (name, category) in indicators.items():
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="1y")

            if hist.empty:
                continue

            close = hist['Close']
            current_price = close.iloc[-1]

            # 소수점 자릿수 결정
            decimals = 4 if category == '환율' else (3 if category == '채권' else 2)

            row = {
                'Category': category,
                'Ticker': ticker,
                'Name': name,
                'Price': fmt(current_price, decimals),
            }

            if len(close) >= 2:
                row['Change(%)'] = fmt((close.iloc[-1] / close.iloc[-2] - 1) * 100)
            if len(close) >= 6:
                row['Return1W(%)'] = fmt((close.iloc[-1] / close.iloc[-6] - 1) * 100)
            if len(close) >= 22:
                row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
            if len(close) >= 66:
                row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
            if len(close) >= 252:
                row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-252] - 1) * 100)
                year_data = close.tail(252)
                row['52wHigh'] = fmt(year_data.max(), decimals)
                row['52wLow'] = fmt(year_data.min(), decimals)
                row['From52wHigh(%)'] = fmt((current_price / year_data.max() - 1) * 100)
                row['From52wLow(%)'] = fmt((current_price / year_data.min() - 1) * 100)

            results.append(row)
        except:
            pass

        time.sleep(0.05)

    # ========================================
    # 2. 계산 지표 (스프레드 등)
    # ========================================
    log("  계산 지표 수집 중...")

    # 10Y-3M 스프레드 (침체 신호)
    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        irx = next((r for r in results if r['Ticker'] == '^IRX'), None)
        if tnx and irx and tnx.get('Price') and irx.get('Price'):
            spread_10y_3m = float(tnx['Price']) - float(irx['Price'])
            results.append({
                'Category': '스프레드',
                'Ticker': '10Y-3M',
                'Name': '10년-3개월 스프레드',
                'Price': fmt(spread_10y_3m, 3),
                'Signal': '역전' if spread_10y_3m < 0 else '정상',
            })
    except:
        pass

    # 10Y-2Y 스프레드 (5년물로 대체)
    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        fvx = next((r for r in results if r['Ticker'] == '^FVX'), None)
        if tnx and fvx and tnx.get('Price') and fvx.get('Price'):
            spread_10y_5y = float(tnx['Price']) - float(fvx['Price'])
            results.append({
                'Category': '스프레드',
                'Ticker': '10Y-5Y',
                'Name': '10년-5년 스프레드',
                'Price': fmt(spread_10y_5y, 3),
            })
    except:
        pass

    # 하이일드 스프레드 (HYG-LQD 차이로 추정)
    try:
        hyg = next((r for r in results if r['Ticker'] == 'HYG'), None)
        lqd = next((r for r in results if r['Ticker'] == 'LQD'), None)
        if hyg and lqd and hyg.get('Return1M(%)') and lqd.get('Return1M(%)'):
            # 하이일드가 투자등급 대비 얼마나 언더퍼폼하는지
            hy_spread_proxy = float(lqd['Return1M(%)']) - float(hyg['Return1M(%)'])
            results.append({
                'Category': '스프레드',
                'Ticker': 'HY-IG',
                'Name': '하이일드 스프레드 (proxy)',
                'Price': fmt(hy_spread_proxy, 2),
                'Description': 'LQD-HYG 1개월 수익률 차이',
            })
    except:
        pass

    # ========================================
    # 3. 심리 지표 (웹 스크래핑)
    # ========================================
    log("  심리 지표 수집 중...")

    # Fear & Greed Index
    fg_data = get_fear_greed_index()
    if fg_data:
        results.append({
            'Category': '심리',
            'Ticker': 'F&G',
            'Name': 'Fear & Greed Index',
            'Price': fg_data.get('value'),
            'Signal': fg_data.get('rating', ''),
            'Previous': fg_data.get('previous_close'),
        })

    # S&P 500 CAPE
    cape = get_cape_ratio()
    if cape:
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'CAPE',
            'Name': 'Shiller CAPE Ratio',
            'Price': cape,
            'Signal': '고평가' if float(cape) > 30 else ('저평가' if float(cape) < 15 else '보통'),
        })

    # S&P 500 Forward PE
    forward_pe = get_sp500_forward_pe()
    if forward_pe:
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'SPY-PE',
            'Name': 'S&P500 PE (trailing)',
            'Price': forward_pe,
        })

    # ★ ISM 제조업 PMI 추가
    ism_pmi = get_ism_pmi()
    if ism_pmi:
        results.append({
            'Category': '경기',
            'Ticker': 'ISM-PMI',
            'Name': 'ISM 제조업 PMI',
            'Price': ism_pmi,
            'Signal': '확장' if float(ism_pmi) > 50 else '수축',
        })

    # ========================================
    # 4. 한국 시장 지표 (pykrx)
    # ========================================
    log("  한국 시장 지표 수집 중...")

    kr_indicators = get_korea_market_indicators()

    if kr_indicators.get('kospi_per'):
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'KOSPI-PER',
            'Name': '코스피 PER',
            'Price': kr_indicators['kospi_per'],
        })

    if kr_indicators.get('kospi_pbr'):
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'KOSPI-PBR',
            'Name': '코스피 PBR',
            'Price': kr_indicators['kospi_pbr'],
        })

    if kr_indicators.get('kospi_div'):
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'KOSPI-DIV',
            'Name': '코스피 배당수익률',
            'Price': kr_indicators['kospi_div'],
        })

    if kr_indicators.get('foreign_net_buy') is not None:
        val = kr_indicators['foreign_net_buy']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-외국인',
            'Name': '외국인 순매수 (5일, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    # ★ 외국인 20일 누적 추가
    if kr_indicators.get('foreign_net_buy_20d') is not None:
        val = kr_indicators['foreign_net_buy_20d']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-외국인-20D',
            'Name': '외국인 순매수 (20일 누적, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    if kr_indicators.get('individual_net_buy') is not None:
        val = kr_indicators['individual_net_buy']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-개인',
            'Name': '개인 순매수 (5일, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    if kr_indicators.get('institution_net_buy') is not None:
        val = kr_indicators['institution_net_buy']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-기관',
            'Name': '기관 순매수 (5일, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    # ========================================
    # 5. 추가 ETF 기반 지표
    # ========================================
    log("  추가 ETF 지표 수집 중...")

    additional_etfs = {
        'IEF': ('미국채 7-10년 ETF', '채권'),
        'TIP': ('물가연동채 ETF', '채권'),
        'EMB': ('신흥국 채권 ETF', '신용'),
        'HYG': ('하이일드 채권 ETF', '신용'),      # ★ 하이일드 스프레드용
        'LQD': ('투자등급 회사채 ETF', '신용'),    # ★ 하이일드 스프레드용
        'GLD': ('금 ETF', '원자재'),
        'USO': ('원유 ETF', '원자재'),
        'VXX': ('VIX 선물 ETF', '변동성'),
        'SVXY': ('VIX 인버스 ETF', '변동성'),
    }

    for ticker, (name, category) in additional_etfs.items():
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="3mo")

            if hist.empty:
                continue

            close = hist['Close']
            current_price = close.iloc[-1]

            row = {
                'Category': category,
                'Ticker': ticker,
                'Name': name,
                'Price': fmt(current_price, 2),
            }

            if len(close) >= 22:
                row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)

            results.append(row)
        except:
            pass

        time.sleep(0.05)

    # 실질금리 추정 (10Y - TIP 수익률)
    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        tip = next((r for r in results if r['Ticker'] == 'TIP'), None)
        if tnx and tip:
            # TIP의 1년 수익률을 기대 인플레이션 proxy로 사용
            tip_return = tip.get('Return1M(%)')
            if tip_return:
                # 단순 추정: 10Y 금리 - (TIP 수익률 * 12 / 100)
                real_rate = float(tnx['Price']) - (float(tip_return) * 12 / 100)
                results.append({
                    'Category': '금리',
                    'Ticker': 'REAL-RATE',
                    'Name': '실질금리 추정',
                    'Price': fmt(real_rate, 3),
                    'Description': '10Y - TIP implied inflation',
                })
    except:
        pass

    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# Excel 저장
# ============================================================
def save_to_excel(data_dict, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    log("\n엑셀 저장 중...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for sheet_name, df in data_dict.items():
        if df is None or df.empty:
            continue
        
        ws = wb.create_sheet(title=sheet_name[:31])
        
        for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.freeze_panes = 'A2'
        log(f"  ✅ {sheet_name}: {len(df)}행")
    
    wb.save(filename)
    log(f"\n💾 Excel 저장: {filename}")

# ============================================================
# JSON 저장
# ============================================================
def save_to_json(data_dict, filename):
    log("\nJSON 저장 중...")
    
    output = {
        'metadata': {
            'generated': datetime.now().isoformat(),
            'date': TODAY,
            'version': 'v2-github-pwa-compatible'  # ← 버전 업데이트
        },
        'data': {}
    }
    
    for sheet_name, df in data_dict.items():
        if df is None or df.empty:
            continue
        
        records = df.replace({np.nan: None}).to_dict(orient='records')
        output['data'][sheet_name] = records
        log(f"  ✅ {sheet_name}: {len(records)}개")
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    log(f"\n💾 JSON 저장: {filename}")
    
    size_mb = os.path.getsize(filename) / (1024 * 1024)
    log(f"   파일 크기: {size_mb:.2f} MB")

# ============================================================
# 메인
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='글로벌 주식/ETF 스크리너 (GitHub Actions)')
    parser.add_argument('--json-only', action='store_true', help='JSON만 출력')
    parser.add_argument('--output-dir', type=str, default='.', help='출력 디렉토리')
    parser.add_argument('--kr-stocks', type=int, default=150, help='한국 주식 수 (기본 150)')
    parser.add_argument('--us-stocks', type=int, default=100, help='미국 주식 수 (기본 100)')
    parser.add_argument('--kr-etfs', type=int, default=200, help='한국 ETF 수 (기본 200)')
    parser.add_argument('--us-etfs', type=int, default=100, help='미국 ETF 수 (기본 100)')
    args = parser.parse_args()
    
    global TOP_N_KR, TOP_N_US, TOP_N_KR_ETF, TOP_N_US_ETF
    TOP_N_KR = args.kr_stocks
    TOP_N_US = args.us_stocks
    TOP_N_KR_ETF = args.kr_etfs
    TOP_N_US_ETF = args.us_etfs
    
    log("=" * 60)
    log("글로벌 주식/ETF 스크리닝 - GitHub Actions v2.5")
    log(f"실행: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"한국 주식: {TOP_N_KR}개 (KOSPI) | 미국 주식: {TOP_N_US}개")
    log(f"한국 ETF: {TOP_N_KR_ETF}개 | 미국 ETF: {TOP_N_US_ETF}개")
    log("=" * 60)
    
    start = time.time()
    
    data = {
        'KR_Stocks': get_korea_stocks(),
        'US_Stocks': get_us_stocks(),
        'KR_ETF': get_korea_etfs(),
        'US_ETF': get_us_etfs(),
        'Market_Indicators': get_market_indicators()
    }
    
    output_dir = args.output_dir
    os.makedirs(output_dir, exist_ok=True)
    
    json_file = os.path.join(output_dir, 'data.json')
    save_to_json(data, json_file)
    
    if not args.json_only:
        excel_file = os.path.join(output_dir, f'global_screening_{TODAY}.xlsx')
        save_to_excel(data, excel_file)
    
    elapsed = (time.time() - start) / 60
    log(f"\n총 소요: {elapsed:.1f}분")
    log("=" * 60)

if __name__ == "__main__":
    main()
