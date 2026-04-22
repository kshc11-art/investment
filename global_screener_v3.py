#!/usr/bin/env python3
"""
글로벌 주식/ETF 스크리닝 - GitHub Actions 버전 v4.1
=============================================================================
v4.1 주요 변경:
  1. 한국 국고채 수익률 추가: 1Y, 3Y, 10Y (네이버금융 / FRED / FDR 멀티소스)
  2. 한국은행 기준금리, CD 91일 금리 추가 (한국 금융 NIM 판단용)
  3. 일본 JGB 10년 수익률 추가 (미국 장기채 환류 리스크 판단용)
  4. 미국채 2년 수익률 추가 → 10Y-2Y 스프레드 계산
  5. 한미 금리차 (US2Y-KR3Y) 추가 → 원화 환율 방향 판단
  6. 한국 국고 10Y-1Y 스프레드 추가 → 한국 수익률 곡선
=============================================================================
v4.0 주요 변경:
  1. US Stocks: 하드코딩 제거 → S&P500 전체 동적 조회 (최대 500개)
  2. US ETF: 하드코딩 제거 → 동적 ETF 리스트 조회 (최대 300개)
  3. KR Stocks: KOSPI + KOSDAQ 확장 (최대 500개, 시총순)
  4. KR ETF: 개수 확대 및 동적 우선
  5. 데이터 검증 강화: 배당률 이상치, 수익률 이상치, 재무 이상치 자동 보정
  6. 거시경제 지표 추가: 기준금리, CPI, 실업률, GDP, M2, 소비자신뢰지수 등
  7. 하드코딩 목록은 최후의 폴백으로만 사용
=============================================================================
v3.0.2 이전:
  - FinanceDataReader 추가, 데이터 소스 우선순위 개선
  - Return1Y 인덱스 버그 수정, 무위험수익률 통일 (4%)
  - SharpeRatio 계산 조건 완화 (252→200일)
  - 기술적 지표 공통 함수, 하드코딩 폴백 (ExpenseRatio 등)

GitHub Actions에서 자동 실행 → JSON 출력 → GitHub Pages에서 PWA가 fetch

설치:
pip install yfinance openpyxl pandas requests beautifulsoup4 lxml numpy pykrx finance-datareader

실행:
python global_screener_v3.py              # Excel + JSON 출력
python global_screener_v3.py --json-only  # JSON만 출력
"""

import warnings
warnings.filterwarnings('ignore')

import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import time
import sys
import re
import json
import argparse

# ★ GitHub Actions 버퍼링 해결: 즉시 출력 함수
def log(msg):
    """즉시 출력되는 로그 함수"""
    print(msg)
    sys.stdout.flush()

log("라이브러리 로딩 중...")

import logging
logging.getLogger('yfinance').setLevel(logging.CRITICAL)
logging.getLogger('urllib3').setLevel(logging.CRITICAL)

try:
    import yfinance as yf
    import requests
    from bs4 import BeautifulSoup
except ImportError as e:
    log("=" * 60)
    log("pip install yfinance openpyxl pandas requests beautifulsoup4 lxml numpy pykrx finance-datareader")
    log("=" * 60)
    sys.exit(1)

# ★ v3.0.2 추가: FinanceDataReader (한국 주식/ETF 최우선)
try:
    import FinanceDataReader as fdr
    FDR_AVAILABLE = True
    log("✅ FinanceDataReader 로드 완료")
except:
    FDR_AVAILABLE = False
    log("⚠️ FinanceDataReader 미설치 - pip install finance-datareader")

# pykrx 선택적
try:
    from pykrx import stock as pykrx_stock
    PYKRX_AVAILABLE = True
    log("✅ pykrx 로드 완료")
except:
    PYKRX_AVAILABLE = False
    log("⚠️ pykrx 미설치 - 한국 주식 일부 데이터 제한")

# ============================================================
# ★ v4.5 추가: Phase B/C 독립실행형 모듈 (optional)
# ============================================================
try:
    from phase_b_financial_events_standalone import integrate_phase_b_standalone
    PHASE_B_AVAILABLE = True
    log("✅ Phase B (DART 재무 이벤트) 모듈 로드")
except ImportError:
    PHASE_B_AVAILABLE = False
    log("ℹ️  Phase B 모듈 없음 (선택적 기능)")

try:
    from phase_c_flow_signals_standalone import integrate_phase_c_standalone
    PHASE_C_AVAILABLE = True
    log("✅ Phase C (KRX 수급 시그널) 모듈 로드")
except ImportError:
    PHASE_C_AVAILABLE = False
    log("ℹ️  Phase C 모듈 없음 (선택적 기능)")

# ============================================================
# ★★★ v3.0 추가: 하드코딩 데이터 ★★★
# ============================================================

# US ETF 비용비율 (2024년 기준, %) - yfinance가 제공하지 않음
US_ETF_EXPENSE = {
    'SPY': 0.0945, 'IVV': 0.03, 'VOO': 0.03, 'VTI': 0.03, 'QQQ': 0.20,
    'DIA': 0.16, 'IWM': 0.19, 'IWF': 0.19, 'IWD': 0.19, 'VUG': 0.04,
    'VTV': 0.04, 'IJH': 0.05, 'IJR': 0.06, 'VB': 0.05, 'VO': 0.04,
    'RSP': 0.20, 'SPLG': 0.02, 'SCHX': 0.03, 'SCHB': 0.03, 'MGK': 0.07,
    'XLK': 0.09, 'XLF': 0.09, 'XLV': 0.09, 'XLE': 0.09, 'XLI': 0.09,
    'XLY': 0.09, 'XLP': 0.09, 'XLU': 0.09, 'XLB': 0.09, 'XLRE': 0.09,
    'VGT': 0.10, 'VFH': 0.10, 'VHT': 0.10, 'VDE': 0.10, 'VIS': 0.10,
    'VCR': 0.10, 'VDC': 0.10, 'VPU': 0.10, 'VAW': 0.10, 'VNQ': 0.12,
    'ARKK': 0.75, 'ARKW': 0.75, 'ARKF': 0.75, 'ARKG': 0.75,
    'SOXX': 0.35, 'SMH': 0.35, 'XBI': 0.35, 'IBB': 0.45,
    'HACK': 0.60, 'BOTZ': 0.68, 'LIT': 0.75, 'TAN': 0.50,
    'ICLN': 0.40, 'PBW': 0.65, 'QCLN': 0.58, 'REMX': 0.75, 'COPX': 0.65,
    'URA': 0.69, 'VYM': 0.06, 'SCHD': 0.06, 'DVY': 0.38, 'HDV': 0.08,
    'SPHD': 0.30, 'SPYD': 0.07, 'VIG': 0.06, 'DGRO': 0.08, 'NOBL': 0.35,
    'BND': 0.03, 'AGG': 0.03, 'TLT': 0.15, 'IEF': 0.15, 'SHY': 0.15,
    'LQD': 0.14, 'HYG': 0.49, 'JNK': 0.40, 'TIP': 0.19,
    'GLD': 0.40, 'IAU': 0.25, 'SLV': 0.50, 'USO': 0.79, 'UNG': 1.35,
    'DBC': 0.85, 'PDBC': 0.59,
    'TQQQ': 0.86, 'SQQQ': 0.86, 'UPRO': 0.91, 'SPXU': 0.91,
    'SOXL': 0.76, 'SOXS': 0.76,
    'VEA': 0.05, 'VWO': 0.08, 'EFA': 0.32, 'EEM': 0.68,
    'IEFA': 0.07, 'IEMG': 0.09, 'VXUS': 0.07, 'ACWI': 0.32
}

# KR ETF 비용비율 (총보수, %) - yfinance가 제공하지 않음
KR_ETF_EXPENSE = {
    '069500': 0.05,   # KODEX 200
    '114800': 0.07,   # KODEX 인버스
    '122630': 0.07,   # KODEX 레버리지
    '229200': 0.25,   # KODEX 코스닥150
    '252670': 0.64,   # KODEX 200선물인버스2X
    '305720': 0.45,   # KODEX 2차전지산업
    '091160': 0.45,   # KODEX 반도체
    '133690': 0.07,   # TIGER 나스닥100
    '143850': 0.07,   # TIGER S&P500
    '192090': 0.55,   # TIGER 차이나CSI300
    '371460': 0.07,   # TIGER 미국나스닥100커버드콜
    '360200': 0.25,   # KINDEX 미국다우존스
    '148020': 0.45,   # KODEX MSCI KOREA
    '153130': 0.24,   # KODEX 단기채권
    '161510': 0.04,   # ARIRANG 고배당주
    '157450': 0.04,   # TIGER 고배당
    '261240': 0.35,   # KODEX USD Futures
    '360750': 0.07,   # TIGER S&P500
    '371450': 0.49,   # TIGER 글로벌클라우드컴퓨팅
    '379800': 0.07,   # KODEX S&P500
    '381170': 0.49,   # TIGER 미국테크TOP10 INDXX
    '395160': 0.45,   # KODEX AI반도체핵심장비
    '461460': 0.10,   # PLUS 10년국채액티브
    '102110': 0.05,   # TIGER 200
    '105190': 0.07,   # KINDEX 200
    '226490': 0.07,   # KODEX 코스피
    '102780': 0.07,   # KODEX 삼성그룹
    '091170': 0.45,   # KODEX 은행
    '091180': 0.45,   # KODEX 자동차
    '117680': 0.45,   # KODEX 건설
    '117700': 0.45,   # KODEX 철강
    '139260': 0.45,   # TIGER 미디어컨텐츠
    '139280': 0.45,   # TIGER 경기방어
    '140700': 0.07,   # KODEX 국채선물
    '148070': 0.50,   # KOSEF 단기자금
    '152500': 0.07,   # KINDEX 레버리지
    '156080': 0.50,   # KODEX MSCI World
    '167860': 0.45,   # KOSEF 미국달러선물
    '182490': 0.07,   # TIGER 나스닥바이오
    '195930': 0.45,   # TIGER 유로스탁스50
    '195980': 0.50,   # ARIRANG 신흥국MSCI
    '200250': 0.07,   # KOSEF 미국S&P500선물
    '210780': 0.07,   # TIGER 코스피고배당
    '211210': 0.07,   # KODEX 종합채권
    '214980': 0.07,   # KODEX 단기채권PLUS
    '217770': 0.50,   # TIGER 원유선물
    '219390': 0.07,   # KINDEX 미국S&P500
    '219480': 0.50,   # KODEX 미국S&P500선물
    '226380': 0.50,   # KINDEX 유로스탁스50
    '227830': 0.07,   # ARIRANG 코스피50
    '228790': 0.07,   # TIGER 화장품
    '228800': 0.07,   # TIGER 여행레저
    '228810': 0.07,   # TIGER 미디어컨텐츠
    '228820': 0.07,   # TIGER 은행
    '233740': 0.50,   # TIGER 원유선물Enhanced
    '233160': 0.50,   # TIGER 코스닥150레버리지
    '238720': 0.07,   # KINDEX 코스닥150
    '243880': 0.07,   # TIGER 200에너지화학
    '243890': 0.07,   # TIGER 200건설
    '244580': 0.50,   # KODEX 골드선물
    '244620': 0.50,   # KODEX 은선물
    '245340': 0.50,   # TIGER 리츠부동산인프라
    '245710': 0.07,   # TIGER 코스닥150
    '251340': 0.50,   # KODEX 선진국MSCI World
    '252400': 0.50,   # KODEX 200동일가중
    '252710': 0.07,   # TIGER 200선물레버리지
    '253150': 0.50,   # ARIRANG 고배당저변동
    '253160': 0.50,   # ARIRANG 스마트베타Quality
    '253240': 0.50,   # KODEX 코스닥150선물인버스
    '261110': 0.07,   # TIGER 코스피대형주
    '261120': 0.07,   # TIGER 코스피중형주
    '266370': 0.50,   # KODEX 인도Nifty50
    '267440': 0.07,   # TIGER 코스닥150IT
    '267490': 0.50,   # KINDEX 인버스
    '267500': 0.50,   # KINDEX 레버리지
    '268280': 0.50,   # KINDEX 200선물레버리지
    '269420': 0.07,   # KODEX S&P글로벌인프라
    '270800': 0.50,   # TIGER 미국채10년선물
    '272220': 0.50,   # TIGER 코스닥150선물인버스
    '272560': 0.50,   # TIGER S&P500선물인버스
    '272580': 0.07,   # TIGER 미국S&P500
    '273130': 0.07,   # KODEX 종합채권(AA-)
    '273140': 0.07,   # KODEX 현금배당성장
    '273210': 0.50,   # ARIRANG 미국S&P500
    '273220': 0.50,   # ARIRANG 미국나스닥100
    '275980': 0.07,   # TIGER 200커버드콜5%OTM
    '276650': 0.07,   # TIGER 은행고배당
    '276990': 0.45,   # KINDEX 일본TOPIX100
    '277630': 0.45,   # TIGER 코스피대형가치
    '277640': 0.45,   # TIGER 코스피대형성장
    '278240': 0.45,   # ARIRANG 스마트베타Momentum
    '278420': 0.07,   # KODEX 코스피대형주
    '278530': 0.45,   # TIGER 200에너지화학레버리지
    '278540': 0.45,   # TIGER 200IT레버리지
    '284430': 0.07,   # KODEX 코스닥150선물레버리지
    '284980': 0.10,   # HANARO 200
    '287300': 0.10,   # HANARO KOSDAQ150
    '287310': 0.07,   # KINDEX 코스피
    '287320': 0.50,   # KINDEX 미국달러선물레버리지
    '287330': 0.50,   # KINDEX 미국달러선물인버스
    '290080': 0.50,   # KODEX 미국채Ultra30년선물
    '292150': 0.50,   # TIGER 미국채30년선물
    '292160': 0.50,   # TIGER 미국채10년선물레버리지
    '292170': 0.50,   # TIGER 미국채10년선물인버스
    '292180': 0.50,   # TIGER 미국채10년선물인버스2X
    '292190': 0.50,   # TIGER 미국채30년선물인버스
    '294400': 0.50,   # KODEX 미국채Ultra10년선물
    '298340': 0.45,   # TIGER 2차전지테마
    '298770': 0.45,   # KODEX 미국채10년선물
    '299660': 0.07,   # TIGER 200산업재
    '300640': 0.50,   # KODEX 코스닥150선물레버리지
    '300950': 0.50,   # TIGER 나스닥바이오텍
    '302190': 0.50,   # TIGER 미국달러선물레버리지
    '304770': 0.45,   # TIGER 게임
    '304780': 0.45,   # TIGER 바이오
    '304940': 0.45,   # KODEX 바이오
    '305080': 0.45,   # TIGER 미디어콘텐츠
    '307510': 0.50,   # ARIRANG ESG종합채권
    '315960': 0.50,   # ARIRANG KS채권혼합
    '319870': 0.50,   # TIGER 200커버드콜ATM
    '322120': 0.50,   # TIGER 통신TOP10
    '322130': 0.45,   # TIGER 의료기기
    '322400': 0.50,   # KODEX 테슬라인컴인버스
    '322410': 0.50,   # KODEX 테슬라인컴레버리지
    '322500': 0.50,   # KODEX 테슬라인컴
    '329750': 0.50,   # TIGER 미국테크TOP10INDXX
    '333940': 0.50,   # ARIRANG S&P500
    '337140': 0.50,   # KODEX 3대농산물선물
    '337160': 0.45,   # KODEX 게임산업
    '352540': 0.45,   # TIGER 헬스케어
    '352560': 0.50,   # TIGER K리츠
    '354350': 0.50,   # KODEX 인버스2X
    '360140': 0.50,   # TIGER AI코리아그로스
    '363570': 0.50,   # KODEX 자동차
    '363580': 0.50,   # KODEX 은행
    '364970': 0.50,   # TIGER KRX바이오K뉴딜
    '364980': 0.50,   # TIGER KRX2차전지K뉴딜
    '365000': 0.50,   # TIGER KRX인터넷K뉴딜
    '365040': 0.50,   # KODEX K뉴딜디지털플러스
    '367380': 0.50,   # TIGER 미국AI테크TOP10
    '368190': 0.50,   # TIGER AI&로봇
    '368590': 0.50,   # KODEX 글로벌리튬
    '371150': 0.50,   # TIGER KRX BBIG K뉴딜
    '371160': 0.50,   # TIGER 차이나전기차레버리지
    '372790': 0.50,   # TIGER 반도체TOP10
    '373490': 0.50,   # TIGER 우주방산
    '375270': 0.50,   # TIGER 코스닥150리밸런싱
    '375720': 0.50,   # TIGER 2차전지테크
    '381180': 0.50,   # TIGER 미국필라델피아반도체
    '385720': 0.50,   # TIGER Fn반도체TOP10
    '385550': 0.50,   # KODEX K-방산
    '385560': 0.50,   # KODEX 글로벌K뉴딜
    '385590': 0.50,   # TIGER K로봇
    '385600': 0.50,   # TIGER K AI반도체핵심장비
    '387270': 0.50,   # TIGER 차이나항셍테크레버리지
    '387280': 0.50,   # TIGER 차이나항셍테크인버스
    '391600': 0.50,   # TIGER AI반도체핵심소재
    '391680': 0.50,   # TIGER AI코리아펀더멘탈
    '394660': 0.50,   # TIGER 글로벌자율주행
    '394670': 0.50,   # TIGER 글로벌BBIG핀테크
    '396070': 0.50,   # KODEX K방산
    '400760': 0.50,   # TIGER 글로벌2차전지TOP10
    '401470': 0.50,   # TIGER 글로벌AI로봇&자율주행
    '402340': 0.50,   # HANARO 글로벌AI에너지
    '404780': 0.50,   # HANARO 글로벌탄소중립
    '411060': 0.50,   # KODEX K게임
    '445280': 0.50,   # TIGER AI BIGTECH 10
    '453810': 0.50,   # TIGER 미국테크TOP10커버드콜
    '458730': 0.50,   # TIGER 미국30년국채스트립액티브
    '459000': 0.50,   # TIGER 미국채30년스트립액티브
    '459100': 0.50,   # KODEX 미국30년국채스트립액티브
    '459200': 0.50,   # TIGER 미국10년국채스트립액티브
    '459580': 0.50,   # TIGER 글로벌온디바이스AI
    '462320': 0.50,   # TIGER 미국배당다우존스
    '464510': 0.50,   # TIGER 미국캐시카우100커버드콜
}

# KR ETF 카테고리
KR_ETF_CATEGORY = {
    '069500': '국내 대형주',
    '114800': '인버스',
    '122630': '레버리지',
    '229200': '코스닥',
    '252670': '인버스 레버리지',
    '305720': '섹터(2차전지)',
    '091160': '섹터(반도체)',
    '133690': '미국 대형주',
    '143850': '미국 대형주',
    '192090': '중국',
    '371460': '커버드콜',
    '360200': '미국 대형주',
    '148020': '국내 전체',
    '153130': '채권(단기)',
    '161510': '배당',
    '157450': '배당',
    '261240': '통화(달러)',
    '360750': '미국 대형주',
    '371450': '테마(클라우드)',
    '379800': '미국 대형주',
    '381170': '미국 테크',
    '395160': '섹터(AI반도체)',
    '461460': '채권(10년)',
    '102110': '국내 대형주',
    '105190': '국내 대형주',
    '226490': '국내 전체',
    '102780': '섹터(삼성그룹)',
    '091170': '섹터(은행)',
    '091180': '섹터(자동차)',
    '117680': '섹터(건설)',
    '117700': '섹터(철강)',
    '139260': '섹터(미디어)',
    '139280': '섹터(경기방어)',
}

# KR ETF 배당수익률 (2024년 기준 추정)
KR_ETF_DIVYIELD = {
    '069500': 1.8,   # KODEX 200
    '114800': 0.0,   # KODEX 인버스
    '122630': 0.0,   # KODEX 레버리지
    '229200': 0.3,   # KODEX 코스닥150
    '252670': 0.0,   # 인버스2X
    '305720': 0.0,   # 2차전지
    '091160': 0.5,   # 반도체
    '133690': 0.4,   # 나스닥100
    '143850': 1.2,   # S&P500
    '192090': 0.8,   # 차이나
    '161510': 4.5,   # 고배당주
    '157450': 3.8,   # TIGER 고배당
    '261240': 0.0,   # 달러선물
    '360750': 1.2,   # S&P500
    '371450': 0.0,   # 클라우드
    '379800': 1.2,   # S&P500
    '381170': 0.2,   # 테크TOP10
    '395160': 0.0,   # AI반도체
    '461460': 3.0,   # 국채
    '371460': 8.0,   # 커버드콜
    '102110': 1.8,   # TIGER 200
    '105190': 1.8,   # KINDEX 200
}

# ============================================================
# 설정
# ============================================================
TODAY = datetime.now().strftime("%Y%m%d")
DATE_1Y_AGO = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")

# 종목 수 설정 - v4.0: 대폭 확대
TOP_N_KR = 500    # 한국 주식: KOSPI + KOSDAQ 최대 500개 (시총순)
TOP_N_US = 500    # 미국 주식: S&P500 전체 + α (동적 조회)
TOP_N_KR_ETF = 300  # 한국 ETF: 300개
TOP_N_US_ETF = 300  # 미국 ETF: 300개 (동적 조회)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
}

# 네이버 차단 여부 (런타임에 판단)
NAVER_AVAILABLE = None
FNGUIDE_AVAILABLE = None  # ★ FnGuide도 차단 감지

# 타임아웃 설정 (GitHub Actions용 단축)
TIMEOUT_SHORT = 3
TIMEOUT_LONG = 5

# 데이터 품질 검증용 필수 컬럼
REQUIRED_COLS_KR_STOCK = ['PER', 'PBR', 'ROE(%)', 'Return60D(%)', 'RSI14', 'Volatility20D']
REQUIRED_COLS_US_STOCK = ['PER', 'PBR', 'ROE(%)', 'Return60D(%)', 'RSI14', 'Volatility20D', 'Beta']
REQUIRED_COLS_ETF = ['ExpenseRatio(%)', 'Return3M(%)', 'Volatility20D']

# ============================================================
# 유틸리티 함수
# ============================================================
def fmt(val, decimals=2):
    """숫자 포맷팅"""
    if val is None or pd.isna(val):
        return None
    try:
        return round(float(val), decimals)
    except:
        return None

def safe_get(d, *keys, default=None):
    """딕셔너리에서 안전하게 값 추출"""
    if not isinstance(d, dict):
        return default
    for k in keys:
        if k in d and d[k] is not None:
            return d[k]
    return default

def fetch_yf_with_retry(ticker, max_retries=2, delay=0.3):
    """yfinance 데이터를 재시도 로직과 함께 가져오기"""
    for attempt in range(max_retries):
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="1y")  # 2y → 1y로 단축
            info = t.info

            # 유효한 데이터인지 확인
            if not hist.empty or (info and len(info) > 5):
                return t, hist, info

            if attempt < max_retries - 1:
                time.sleep(delay)
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(delay)
            else:
                pass

    return None, pd.DataFrame(), {}

def try_multiple_tickers(ticker_variants, max_retries=2):
    """여러 티커 형식을 시도하여 가장 좋은 결과 반환"""
    best_result = (None, pd.DataFrame(), {})
    best_score = 0

    for ticker in ticker_variants:
        t, hist, info = fetch_yf_with_retry(ticker, max_retries=max_retries)

        # 점수 계산: hist 길이 + info 키 수
        score = len(hist) + len(info) if info else len(hist)

        if score > best_score:
            best_score = score
            best_result = (t, hist, info)

            # 충분히 좋은 결과면 조기 종료
            if len(hist) > 200 and len(info) > 20:
                break

    return best_result

def fill_missing_from_info(row, info, field_mapping):
    """info 딕셔너리에서 결측값 채우기"""
    for row_field, (info_keys, multiplier, decimals) in field_mapping.items():
        if row.get(row_field) is None:
            val = safe_get(info, *info_keys) if isinstance(info_keys, tuple) else safe_get(info, info_keys)
            if val is not None:
                if multiplier != 1:
                    val = val * multiplier
                row[row_field] = fmt(val, decimals)
    return row

# ============================================================
# 네이버 금융 스크래핑
# ============================================================
def fetch_naver(url, timeout=TIMEOUT_SHORT):
    """네이버 URL 요청"""
    global NAVER_AVAILABLE
    if NAVER_AVAILABLE is False:
        return None
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        if resp.status_code == 403:
            NAVER_AVAILABLE = False
            log("  ⚠️ 네이버 금융 접근 차단됨 (403)")
            return None
        if resp.status_code == 200:
            NAVER_AVAILABLE = True
            return resp
        return None
    except:
        return None

def get_naver_stock_list(market='KOSPI', max_pages=10):
    """네이버에서 시가총액 순위 종목 리스트"""
    global NAVER_AVAILABLE
    stocks = []
    sosok = '0' if market == 'KOSPI' else '1'
    
    log(f"  네이버 {market} 시총 순위 로드 시도...")
    
    for page in range(1, max_pages + 1):
        url = f"https://finance.naver.com/sise/sise_market_sum.naver?sosok={sosok}&page={page}"
        resp = fetch_naver(url)
        
        if not resp:
            if NAVER_AVAILABLE is False:
                return []
            continue
        
        soup = BeautifulSoup(resp.text, 'lxml')
        table = soup.select_one('table.type_2')
        
        if not table:
            continue
        
        for tr in table.select('tr'):
            tds = tr.select('td')
            if len(tds) < 10:
                continue
            
            try:
                link = tds[1].select_one('a')
                if not link:
                    continue
                
                href = link.get('href', '')
                code_match = re.search(r'code=(\d{6})', href)
                if not code_match:
                    continue
                
                code = code_match.group(1)
                name = link.text.strip()
                
                price_text = tds[2].text.strip().replace(',', '')
                price = int(price_text) if price_text.isdigit() else None
                
                change_pct = None
                if len(tds) > 4:
                    try:
                        pct_text = tds[4].text.strip().replace('%', '').replace('+', '')
                        change_pct = float(pct_text) if pct_text else None
                    except:
                        pass
                
                market_cap = None
                if len(tds) > 6:
                    try:
                        cap_text = tds[6].text.strip().replace(',', '')
                        market_cap = int(cap_text) if cap_text.isdigit() else None
                    except:
                        pass
                
                stocks.append({
                    'code': code, 'name': name, 'price': price,
                    'change_pct': change_pct, 'market_cap': market_cap, 'market': market
                })
            except:
                continue
        
        time.sleep(0.2)
    
    log(f"  네이버 {market}: {len(stocks)}개 로드")
    return stocks

def get_naver_stock_detail(code):
    """네이버에서 종목 상세 정보 - 확장 버전"""
    data = {}

    # 1. 메인 페이지에서 기본 정보
    url = f"https://finance.naver.com/item/main.naver?code={code}"
    resp = fetch_naver(url)

    if resp:
        try:
            soup = BeautifulSoup(resp.text, 'lxml')

            # PER, PBR, 배당수익률
            for em in soup.select('em'):
                em_id = em.get('id', '')
                try:
                    if em_id == '_per':
                        val = float(em.text.replace(',', ''))
                        if 0 < val < 1000:
                            data['per'] = val
                    elif em_id == '_pbr':
                        val = float(em.text.replace(',', ''))
                        if 0 < val < 100:
                            data['pbr'] = val
                    elif em_id == '_dvr':
                        val = float(em.text.replace(',', ''))
                        if 0 <= val < 50:
                            data['div_yield'] = val
                    elif em_id == '_eps':
                        val = float(em.text.replace(',', ''))
                        data['eps'] = val
                    elif em_id == '_bps':
                        val = float(em.text.replace(',', ''))
                        data['bps'] = val
                except:
                    pass

            # 현재가
            try:
                price_elem = soup.select_one('p.no_today span.blind')
                if price_elem:
                    data['price'] = int(price_elem.text.replace(',', ''))
            except:
                pass

            # 시가총액
            try:
                for em in soup.select('em#_market_sum'):
                    cap_text = em.text.strip().replace(',', '').replace('조', '').replace('억', '')
                    # 시가총액은 억 단위로 저장
                    if '조' in em.text:
                        data['market_cap'] = int(float(cap_text) * 10000)
                    else:
                        data['market_cap'] = int(cap_text)
            except:
                pass

            # 외국인 비율
            try:
                for table in soup.select('table'):
                    text = table.get_text()
                    if '외국인' in text:
                        match = re.search(r'외국인[^%\d]*([\d.]+)\s*%', text)
                        if match:
                            val = float(match.group(1))
                            if 0 <= val <= 100:
                                data['foreign_ratio'] = val
                                break
            except:
                pass

            # 52주 고저
            try:
                for table in soup.select('table'):
                    for tr in table.select('tr'):
                        th = tr.select_one('th')
                        td = tr.select_one('td')
                        if th and td:
                            th_text = th.get_text().strip()
                            td_text = td.get_text().strip().replace(',', '')
                            if '52주최고' in th_text or '52주 최고' in th_text:
                                match = re.search(r'(\d+)', td_text)
                                if match:
                                    data['high_52w'] = int(match.group(1))
                            elif '52주최저' in th_text or '52주 최저' in th_text:
                                match = re.search(r'(\d+)', td_text)
                                if match:
                                    data['low_52w'] = int(match.group(1))
            except:
                pass

            # 업종(섹터)
            try:
                for a in soup.select('a[href*="upjong"]'):
                    sector_text = a.get_text().strip()
                    if sector_text and len(sector_text) > 1:
                        data['sector'] = sector_text
                        break
            except:
                pass

        except:
            pass

    # 2. 투자지표 페이지에서 ROE, ROA 등 가져오기 (sleep 제거)
    url2 = f"https://finance.naver.com/item/coinfo.naver?code={code}&target=finsum_more"
    resp2 = fetch_naver(url2)

    if resp2:
        try:
            soup2 = BeautifulSoup(resp2.text, 'lxml')

            # iframe 내용이 있을 수 있으므로 테이블에서 직접 추출
            for table in soup2.select('table'):
                rows = table.select('tr')
                for row in rows:
                    cells = row.select('td, th')
                    if len(cells) >= 2:
                        label = cells[0].get_text().strip()

                        # 가장 최근 값 (보통 마지막 td)
                        for cell in reversed(cells[1:]):
                            val_text = cell.get_text().strip().replace(',', '').replace('%', '')
                            if val_text and val_text != '-' and val_text != 'N/A':
                                try:
                                    val = float(val_text)
                                    if 'ROE' in label and 'roe' not in data:
                                        if -100 < val < 200:
                                            data['roe'] = val
                                    elif 'ROA' in label and 'roa' not in data:
                                        if -100 < val < 100:
                                            data['roa'] = val
                                    elif '영업이익률' in label and 'op_margin' not in data:
                                        if -100 < val < 100:
                                            data['op_margin'] = val
                                    elif '순이익률' in label and 'net_margin' not in data:
                                        if -100 < val < 100:
                                            data['net_margin'] = val
                                    elif '부채비율' in label and 'debt_ratio' not in data:
                                        if 0 <= val < 1000:
                                            data['debt_ratio'] = val
                                    elif '유동비율' in label and 'current_ratio' not in data:
                                        if 0 < val < 1000:
                                            data['current_ratio'] = val / 100  # 백분율 → 배수
                                    elif '매출액증가율' in label and 'revenue_growth' not in data:
                                        if -100 < val < 500:
                                            data['revenue_growth'] = val
                                    elif '영업이익증가율' in label and 'op_growth' not in data:
                                        if -200 < val < 1000:
                                            data['op_growth'] = val
                                    break
                                except:
                                    pass
        except:
            pass

    return data

# ============================================================
# FnGuide 스크래핑
# ============================================================
def get_fnguide_data(code):
    """FnGuide에서 재무 데이터 가져오기"""
    global FNGUIDE_AVAILABLE
    
    # 이미 차단된 경우 스킵
    if FNGUIDE_AVAILABLE is False:
        return {}
    
    data = {}
    url = f"https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A{code}"

    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_SHORT)
        if resp.status_code == 403:
            FNGUIDE_AVAILABLE = False
            log("  ⚠️ FnGuide 접근 차단됨")
            return data
        if resp.status_code != 200:
            return data
        
        FNGUIDE_AVAILABLE = True
        soup = BeautifulSoup(resp.text, 'lxml')

        # 시가총액, PER, PBR 등 기본 정보
        try:
            for table in soup.select('table.us_table_ty1'):
                for tr in table.select('tr'):
                    tds = tr.select('td')
                    ths = tr.select('th')
                    if len(ths) >= 1 and len(tds) >= 1:
                        for i, th in enumerate(ths):
                            label = th.get_text().strip()
                            if i < len(tds):
                                val_text = tds[i].get_text().strip().replace(',', '').replace('%', '')
                                if val_text and val_text != '-':
                                    try:
                                        val = float(val_text)
                                        if 'PER' in label and 'per' not in data:
                                            if 0 < val < 1000:
                                                data['per'] = val
                                        elif 'PBR' in label and 'pbr' not in data:
                                            if 0 < val < 100:
                                                data['pbr'] = val
                                        elif 'ROE' in label and 'roe' not in data:
                                            if -100 < val < 200:
                                                data['roe'] = val
                                        elif 'ROA' in label and 'roa' not in data:
                                            if -100 < val < 100:
                                                data['roa'] = val
                                    except:
                                        pass
        except:
            pass

        # 재무비율 테이블에서 추가 정보
        try:
            for table in soup.select('table'):
                text = table.get_text()
                if '영업이익률' in text or '부채비율' in text:
                    for tr in table.select('tr'):
                        cells = tr.select('td, th')
                        if len(cells) >= 2:
                            label = cells[0].get_text().strip()

                            # 최신 값 추출 (마지막 유효한 값)
                            for cell in reversed(cells[1:]):
                                val_text = cell.get_text().strip().replace(',', '').replace('%', '')
                                if val_text and val_text != '-' and val_text != 'N/A':
                                    try:
                                        val = float(val_text)
                                        if '영업이익률' in label and 'op_margin' not in data:
                                            data['op_margin'] = val
                                        elif '순이익률' in label and 'net_margin' not in data:
                                            data['net_margin'] = val
                                        elif '부채비율' in label and 'debt_ratio' not in data:
                                            if 0 <= val < 1000:
                                                data['debt_ratio'] = val
                                        elif '유동비율' in label and 'current_ratio' not in data:
                                            if val > 0:
                                                data['current_ratio'] = val / 100
                                        elif '매출액증가율' in label and 'revenue_growth' not in data:
                                            data['revenue_growth'] = val
                                        elif 'EPS' in label and 'eps' not in data:
                                            data['eps'] = val
                                        elif 'BPS' in label and 'bps' not in data:
                                            data['bps'] = val
                                        break
                                    except:
                                        pass
        except:
            pass

    except:
        pass

    return data

def get_naver_etf_list(max_pages=5):
    """네이버에서 ETF 리스트"""
    etfs = []
    log("  네이버 ETF 리스트 로드 시도...")
    
    for page in range(1, max_pages + 1):
        url = f"https://finance.naver.com/sise/etf.naver?page={page}"
        resp = fetch_naver(url)
        
        if not resp:
            if NAVER_AVAILABLE is False:
                return []
            continue
        
        soup = BeautifulSoup(resp.text, 'lxml')
        
        for table in soup.select('table.type_1, table.type_2'):
            for tr in table.select('tr'):
                tds = tr.select('td')
                if len(tds) < 5:
                    continue
                
                try:
                    link = tds[0].select_one('a')
                    if not link:
                        continue
                    
                    href = link.get('href', '')
                    code_match = re.search(r'code=(\d{6})', href)
                    if not code_match:
                        continue
                    
                    code = code_match.group(1)
                    name = link.text.strip()
                    
                    price = None
                    if len(tds) > 1:
                        price_text = tds[1].text.strip().replace(',', '')
                        price = int(price_text) if price_text.isdigit() else None
                    
                    etfs.append({'code': code, 'name': name, 'price': price})
                except:
                    continue
        
        time.sleep(0.1)
    
    log(f"  네이버 ETF: {len(etfs)}개 로드")
    return etfs

def get_krx_etf_data():
    """KRX에서 ETF 전종목 데이터 직접 수집 (가장 정확)"""
    etf_data = {}
    
    try:
        log("  KRX에서 ETF 전종목 로드 중...")
        
        # 1. ETF 전종목 시세
        gen_otp_url = 'http://data.krx.co.kr/comm/fileDn/GenerateOTP/generate.cmd'
        gen_otp_data = {
            'locale': 'ko_KR',
            'mktId': 'ETF',
            'trdDd': datetime.now().strftime('%Y%m%d'),
            'share': '1',
            'money': '1',
            'csvxls_isNo': 'false',
            'name': 'fileDown',
            'url': 'dbms/MDC/STAT/standard/MDCSTAT04301'
        }
        
        otp_resp = requests.post(gen_otp_url, data=gen_otp_data, headers=HEADERS, timeout=10)
        otp = otp_resp.text
        
        down_url = 'http://data.krx.co.kr/comm/fileDn/download_csv/download.cmd'
        down_resp = requests.post(
            down_url, 
            data={'code': otp}, 
            headers={**HEADERS, 'Referer': gen_otp_url},
            timeout=30
        )
        
        from io import StringIO
        csv_text = down_resp.content.decode('euc-kr', errors='ignore')
        df = pd.read_csv(StringIO(csv_text))
        
        if not df.empty and len(df) > 50:
            log(f"  KRX: {len(df)}개 ETF 로드")
            
            for _, row in df.iterrows():
                code = str(row.get('종목코드', '')).strip()
                if not code or len(code) != 6:
                    continue
                    
                etf_data[code] = {
                    'name': str(row.get('종목명', '')).strip(),
                    'price': row.get('종가', row.get('현재가', None)),
                    'nav': row.get('NAV', row.get('순자산가치', None)),
                    'volume': row.get('거래량', None),
                }
            
            return etf_data
                
    except Exception as e:
        log(f"  ⚠️ KRX 크롤링 실패: {e}")
    
    return {}

# ============================================================
# 기술적 지표 계산
# ============================================================
def calc_rsi(prices, period=14):
    try:
        delta = prices.diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=period).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=period).mean()
        rs = gain / loss
        rsi = 100 - (100 / (1 + rs))
        return rsi.iloc[-1]
    except:
        return None

def calc_volatility(prices, period=20):
    try:
        returns = prices.pct_change().dropna()
        if len(returns) < period:
            return None
        vol = returns.rolling(window=period).std().iloc[-1]
        return vol * np.sqrt(252) * 100
    except:
        return None

def calc_bollinger_position(prices, period=20):
    try:
        ma = prices.rolling(window=period).mean()
        std = prices.rolling(window=period).std()
        upper = ma + 2 * std
        lower = ma - 2 * std
        current = prices.iloc[-1]
        position = (current - lower.iloc[-1]) / (upper.iloc[-1] - lower.iloc[-1]) * 100
        return max(0, min(100, position))
    except:
        return None

def calc_max_drawdown(prices):
    try:
        peak = prices.expanding().max()
        drawdown = (prices - peak) / peak * 100
        return drawdown.min()
    except:
        return None

# ★★★ v3.0 수정: SharpeRatio 계산 조건 완화 ★★★
def calc_sharpe_ratio(prices, min_period=200, risk_free=0.04):
    """
    SharpeRatio 계산 - v3.0 수정
    
    변경사항:
    - 기존: period=252 (252일 미만이면 None)
    - 수정: min_period=200 (200일 이상이면 계산)
    - v3.0.1: risk_free 0.02 → 0.04 (4%)로 통일
    
    이유: yfinance period="1y"가 실제로 약 250-251일 반환
    """
    try:
        returns = prices.pct_change().dropna()
        if len(returns) < min_period:  # ★ 252 → 200으로 완화
            return None
        mean_return = returns.mean() * 252
        std_return = returns.std() * np.sqrt(252)
        if std_return == 0:
            return None
        return (mean_return - risk_free) / std_return
    except:
        return None

# ★★★ v3.0 추가: 기존 데이터 기반 SharpeRatio 계산 ★★★
def calc_sharpe_from_existing(return_pct, volatility_pct, period_days=120, risk_free_rate=4.0):
    """
    이미 수집된 수익률/변동성으로 SharpeRatio 계산
    
    Parameters:
    - return_pct: Return120D(%) 또는 Return6M(%)
    - volatility_pct: Volatility20D (일간 변동성 연환산 %)
    - risk_free_rate: 무위험 수익률 (연 4% 가정)
    """
    try:
        if return_pct is None or volatility_pct is None:
            return None
        if pd.isna(return_pct) or pd.isna(volatility_pct):
            return None
        if volatility_pct == 0:
            return None
        
        # 연환산 수익률
        annual_return = float(return_pct) * (252 / period_days)
        
        # 변동성은 이미 연환산된 값 (Volatility20D)
        annual_vol = float(volatility_pct)
        
        # Sharpe Ratio
        sharpe = (annual_return - risk_free_rate) / annual_vol
        return fmt(sharpe, 3)
    except:
        return None

# ============================================================
# 데이터 검증
# ============================================================
def validate_foreign_ratio(value):
    if value is None:
        return None
    try:
        v = float(value)
        if v < 0 or v > 100:
            return None
        if v > 70:
            return None
        return v
    except:
        return None

def is_etf_stock(name, code=None):
    if not name:
        return False
    name_upper = name.upper()
    etf_brands = ['KODEX', 'TIGER', 'KBSTAR', 'ARIRANG', 'KOSEF', 'KINDEX',
                  'HANARO', 'SOL', 'ACE', 'TIMEFOLIO', 'WOORI', 'VITA',
                  'FOCUS', 'MASTER', 'TREX', 'SMART', 'PLUS', 'RISE']
    for brand in etf_brands:
        if brand.upper() in name_upper:
            return True
    etf_keywords = ['ETF', 'ETN', '레버리지', '인버스', '선물', '지수', '액티브', '채권']
    for kw in etf_keywords:
        if kw in name or kw.upper() in name_upper:
            return True
    return False

def calc_data_quality_score(row, required_cols, data_type='stock'):
    valid_count = 0
    for col in required_cols:
        if col in row and row[col] is not None:
            valid_count += 1
    quality = (valid_count / len(required_cols)) * 100 if required_cols else 100
    row['DataQuality(%)'] = fmt(quality, 0)
    return row

# ============================================================
# ★★★ v4.0 추가: 데이터 검증 및 이상치 보정 ★★★
# ============================================================
def validate_and_clean_row(row, data_type='stock'):
    """
    수집된 데이터의 이상치를 감지하고 보정
    - 배당수익률: yfinance가 소수(0.02) 대신 퍼센트(2.0)를 반환하는 경우 대응
      실제 200%는 비현실적이므로 100으로 나눔 (200→2.0)
    - 수익률: 비현실적 수준 필터
    - 재무지표: 범위 검증
    """
    # 배당수익률 보정: yfinance는 간혹 이미 %인 값을 다시 *100 하는 버그 있음
    div_yield = row.get('DivYield(%)')
    if div_yield is not None:
        try:
            dv = float(div_yield)
            if dv > 30:  # 배당률 30% 초과는 비현실적 → 100으로 나눔
                row['DivYield(%)'] = fmt(dv / 100, 2)
            elif dv < 0:
                row['DivYield(%)'] = None
        except:
            pass

    # PayoutRatio 보정
    payout = row.get('PayoutRatio(%)')
    if payout is not None:
        try:
            pv = float(payout)
            if pv > 200:  # 200% 초과 배당성향은 비현실적
                row['PayoutRatio(%)'] = None
            elif pv < 0:
                row['PayoutRatio(%)'] = None
        except:
            pass

    # PER 이상치 제거
    per = row.get('PER')
    if per is not None:
        try:
            pv = float(per)
            if pv > 500 or pv < 0:  # 음수 PER 또는 500배 초과
                row['PER'] = None
        except:
            pass

    # PBR 이상치 제거
    pbr = row.get('PBR')
    if pbr is not None:
        try:
            pv = float(pbr)
            if pv > 100 or pv < 0:
                row['PBR'] = None
        except:
            pass

    # ROE/ROA 범위 검증
    for field in ['ROE(%)', 'ROA(%)']:
        val = row.get(field)
        if val is not None:
            try:
                v = float(val)
                if v > 300 or v < -200:  # 극단치 제거
                    row[field] = None
            except:
                pass

    # 수익률 이상치 검증 (1일, 1주 수익률)
    for field in ['Return1D(%)', 'Return1W(%)']:
        val = row.get(field)
        if val is not None:
            try:
                v = float(val)
                if abs(v) > 50:  # 1일/1주 50% 이상 변동은 이상치 가능성
                    row[field] = None
            except:
                pass

    # 장기 수익률 이상치 (1년 수익률 1000% 초과)
    for field in ['Return1Y(%)', 'Return250D(%)']:
        val = row.get(field)
        if val is not None:
            try:
                v = float(val)
                if abs(v) > 1000:
                    row[field] = None
            except:
                pass

    # Beta 범위 검증
    beta = row.get('Beta')
    if beta is not None:
        try:
            bv = float(beta)
            if bv > 5 or bv < -3:
                row['Beta'] = None
        except:
            pass

    # ExpenseRatio 검증 (ETF)
    expense = row.get('ExpenseRatio(%)')
    if expense is not None:
        try:
            ev = float(expense)
            if ev > 5 or ev < 0:  # 5% 초과 비용비율 비현실적
                row['ExpenseRatio(%)'] = None
        except:
            pass

    return row

# ============================================================
# ★★★ v3.0 추가: 기술적 지표 공통 계산 함수 ★★★
# ============================================================
def add_technical_indicators(row, close, include_ma60_120=False):
    """
    기술적 지표를 row에 추가하는 공통 함수
    
    Parameters:
    - row: 데이터 딕셔너리
    - close: pandas Series (종가)
    - include_ma60_120: MA60/MA120 포함 여부 (US_Stocks용)
    """
    if close is None or len(close) < 20:
        return row
    
    price = close.iloc[-1]
    
    # 수익률 계산 - ★ v3.0: 조건 완화 (252 → 245)
    if len(close) >= 2:
        row['Return1D(%)'] = fmt((close.iloc[-1] / close.iloc[-2] - 1) * 100)
    if len(close) >= 6:
        row['Return1W(%)'] = fmt((close.iloc[-1] / close.iloc[-6] - 1) * 100)
    if len(close) >= 22:
        row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
    if len(close) >= 66:
        row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
        row['Return60D(%)'] = row['Return3M(%)']  # PWA 호환
    if len(close) >= 132:
        row['Return6M(%)'] = fmt((close.iloc[-1] / close.iloc[-132] - 1) * 100)
        row['Return120D(%)'] = row['Return6M(%)']  # PWA 호환
    
    # ★ v3.0 수정: Return1Y 조건 완화 (252 → 245)
    # ★ v3.0.1 버그 수정: year_ago_idx 계산 (len-1 → len)
    if len(close) >= 245:
        year_ago_idx = min(len(close), 252)  # ★ 수정: -1 제거
        row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-year_ago_idx] - 1) * 100)
        row['Return250D(%)'] = row['Return1Y(%)']  # PWA 호환
    
    # 이동평균
    row['MA20'] = fmt(close.rolling(20).mean().iloc[-1])
    row['MA50'] = fmt(close.rolling(50).mean().iloc[-1])
    row['MA200'] = fmt(close.rolling(200).mean().iloc[-1]) if len(close) >= 200 else None
    
    # ★ v3.0 추가: MA60, MA120 (US_Stocks용)
    if include_ma60_120:
        if len(close) >= 60:
            row['MA60'] = fmt(close.rolling(60).mean().iloc[-1])
        if len(close) >= 120:
            row['MA120'] = fmt(close.rolling(120).mean().iloc[-1])
    
    # vs_MA 계산
    if row.get('MA20'): row['vs_MA20(%)'] = fmt((price / row['MA20'] - 1) * 100)
    if row.get('MA50'): row['vs_MA50(%)'] = fmt((price / row['MA50'] - 1) * 100)
    if row.get('MA200'): row['vs_MA200(%)'] = fmt((price / row['MA200'] - 1) * 100)
    
    # ★ v3.0 추가: vs_MA60, vs_MA120
    if include_ma60_120:
        if row.get('MA60'): row['vs_MA60(%)'] = fmt((price / row['MA60'] - 1) * 100)
        if row.get('MA120'): row['vs_MA120(%)'] = fmt((price / row['MA120'] - 1) * 100)
    
    # RSI, BB
    row['RSI14'] = fmt(calc_rsi(close, 14))
    row['BB_Position'] = fmt(calc_bollinger_position(close, 20))
    
    # 52주 고저
    if len(close) >= 245:  # ★ v3.0: 252 → 245
        year_data = close.tail(min(len(close), 252))
        row['52wHigh'] = fmt(year_data.max())
        row['52wLow'] = fmt(year_data.min())
        row['From52wHigh(%)'] = fmt((price / year_data.max() - 1) * 100)
        row['From52wLow(%)'] = fmt((price / year_data.min() - 1) * 100)
    
    # 변동성
    row['Volatility20D'] = fmt(calc_volatility(close, 20))
    row['Volatility60D'] = fmt(calc_volatility(close, 60))
    
    # MaxDrawdown
    row['MaxDrawdown(%)'] = fmt(calc_max_drawdown(close))
    
    # ★ v3.0 수정: SharpeRatio (조건 완화)
    row['SharpeRatio'] = fmt(calc_sharpe_ratio(close))
    
    # ★ v3.0 추가: SharpeRatio 폴백 (가격 데이터 부족 시 기존 데이터로 계산)
    if row.get('SharpeRatio') is None:
        row['SharpeRatio'] = calc_sharpe_from_existing(
            row.get('Return120D(%)'), 
            row.get('Volatility20D')
        )
    
    return row

# ============================================================
# ★ v3.0.2 추가: FinanceDataReader 데이터 수집 함수
# ============================================================
def fetch_fdr_stock_data(ticker, start_date=None):
    """FinanceDataReader에서 주식 데이터 가져오기"""
    if not FDR_AVAILABLE:
        return None, pd.DataFrame()
    
    try:
        if start_date is None:
            start_date = (datetime.now() - timedelta(days=400)).strftime('%Y-%m-%d')
        
        df = fdr.DataReader(ticker, start_date)
        if df is not None and not df.empty:
            return df, True
    except:
        pass
    return None, False

def fetch_fdr_stock_list(market='KOSPI'):
    """FinanceDataReader에서 종목 리스트 가져오기"""
    if not FDR_AVAILABLE:
        return []
    
    try:
        stocks = fdr.StockListing(market)
        if stocks is not None and not stocks.empty:
            result = []
            for _, row in stocks.iterrows():
                code = str(row.get('Code', row.get('Symbol', ''))).zfill(6)
                name = row.get('Name', row.get('종목명', ''))
                if code and name:
                    result.append((code, name, market))
            return result
    except:
        pass
    return []

def fetch_fdr_etf_list():
    """FinanceDataReader에서 ETF 리스트 가져오기"""
    if not FDR_AVAILABLE:
        return []
    
    try:
        etfs = fdr.StockListing('ETF/KR')
        if etfs is not None and not etfs.empty:
            result = []
            for _, row in etfs.iterrows():
                code = str(row.get('Code', row.get('Symbol', ''))).zfill(6)
                name = row.get('Name', row.get('종목명', ''))
                if code and name:
                    result.append(code)
            return result
    except:
        pass
    return []

# ============================================================
# 한국 주식 수집
# ============================================================
def get_korea_stocks():
    """한국 주식 데이터 수집 - v4.0: KOSPI + KOSDAQ, 시총순 최대 500개"""
    log("\n[1/6] 한국 주식 수집 중...")

    all_tickers = []

    # ★ v4.0: KOSPI + KOSDAQ 모두 수집
    if FDR_AVAILABLE:
        try:
            log("  FinanceDataReader에서 종목 리스트 로드 중 (KOSPI + KOSDAQ)...")
            for market in ['KOSPI', 'KOSDAQ']:
                fdr_stocks = fetch_fdr_stock_list(market)
                for ticker, name, mkt in fdr_stocks:
                    if not is_etf_stock(name, ticker):
                        all_tickers.append((ticker, name, mkt))

            if all_tickers:
                log(f"  FinanceDataReader: {len(all_tickers)}개 종목 로드 (KOSPI+KOSDAQ)")
        except Exception as e:
            log(f"  ⚠️ FinanceDataReader 실패: {e}")
            all_tickers = []

    # 2순위: pykrx (KOSPI + KOSDAQ)
    if not all_tickers and PYKRX_AVAILABLE:
        try:
            log("  pykrx에서 종목 리스트 로드 중 (KOSPI + KOSDAQ)...")
            for market in ['KOSPI', 'KOSDAQ']:
                try:
                    market_tickers = pykrx_stock.get_market_ticker_list(market=market)
                    for ticker in market_tickers:
                        try:
                            name = pykrx_stock.get_market_ticker_name(ticker)
                        except:
                            name = ticker
                        if not is_etf_stock(name, ticker):
                            all_tickers.append((ticker, name, market))
                except:
                    pass

            log(f"  pykrx: {len(all_tickers)}개 종목 로드 (KOSPI+KOSDAQ)")
        except Exception as e:
            log(f"  ⚠️ pykrx 실패: {e}")
            all_tickers = []
    
    # 3순위: 네이버
    if not all_tickers:
        log("  네이버에서 종목 리스트 로드 시도...")
        kospi_stocks = get_naver_stock_list('KOSPI', max_pages=10 if TOP_N_KR is None else 3)
        
        for stock in kospi_stocks:
            if not is_etf_stock(stock['name'], stock['code']):
                all_tickers.append((stock['code'], stock['name'], 'KOSPI'))
    
    if not all_tickers:
        log("  ❌ 종목 리스트를 가져올 수 없음")
        return pd.DataFrame()
    
    if TOP_N_KR:
        all_tickers = all_tickers[:TOP_N_KR]
    
    log(f"  대상: {len(all_tickers)}개")
    
    results = []
    start_time = time.time()
    
    for i, (ticker, name, market) in enumerate(all_tickers):
        try:
            row = {'Code': ticker, 'Name': name, 'Market': market}

            # ========================================
            # 1. 네이버 (주요 소스)
            # ========================================
            naver_data = {}
            if NAVER_AVAILABLE is not False:
                naver_data = get_naver_stock_detail(ticker)
                if naver_data:
                    if naver_data.get('price'):
                        row['Price'] = naver_data['price']
                    if naver_data.get('market_cap'):
                        row['MarketCap(억)'] = naver_data['market_cap']
                    if naver_data.get('sector'):
                        row['Sector'] = naver_data['sector']
                    if naver_data.get('per'):
                        row['PER'] = fmt(naver_data['per'])
                    if naver_data.get('pbr'):
                        row['PBR'] = fmt(naver_data['pbr'])
                    if naver_data.get('eps'):
                        row['EPS'] = fmt(naver_data['eps'], 0)
                    if naver_data.get('bps'):
                        row['BPS'] = fmt(naver_data['bps'], 0)
                    if naver_data.get('roe'):
                        row['ROE(%)'] = fmt(naver_data['roe'])
                    if naver_data.get('roa'):
                        row['ROA(%)'] = fmt(naver_data['roa'])
                    if naver_data.get('op_margin'):
                        row['OpMargin(%)'] = fmt(naver_data['op_margin'])
                    if naver_data.get('net_margin'):
                        row['NetMargin(%)'] = fmt(naver_data['net_margin'])
                    if naver_data.get('revenue_growth'):
                        row['RevenueGrowth(%)'] = fmt(naver_data['revenue_growth'])
                    if naver_data.get('op_growth'):
                        row['EarningsGrowth(%)'] = fmt(naver_data['op_growth'])
                    if naver_data.get('debt_ratio'):
                        row['DebtRatio(%)'] = fmt(naver_data['debt_ratio'])
                    if naver_data.get('current_ratio'):
                        row['CurrentRatio'] = fmt(naver_data['current_ratio'])
                    if naver_data.get('foreign_ratio'):
                        row['ForeignRatio(%)'] = fmt(naver_data['foreign_ratio'])
                    if naver_data.get('div_yield'):
                        row['DivYield(%)'] = fmt(naver_data['div_yield'])
                    if naver_data.get('high_52w'):
                        row['52wHigh'] = naver_data['high_52w']
                    if naver_data.get('low_52w'):
                        row['52wLow'] = naver_data['low_52w']

            # ========================================
            # 2. FnGuide (결측값 폴백)
            # ========================================
            need_fnguide = (
                row.get('ROE(%)') is None or
                row.get('ROA(%)') is None or
                row.get('OpMargin(%)') is None or
                row.get('DebtRatio(%)') is None
            )

            if need_fnguide:
                fnguide_data = get_fnguide_data(ticker)
                if fnguide_data:
                    if row.get('PER') is None and fnguide_data.get('per'):
                        row['PER'] = fmt(fnguide_data['per'])
                    if row.get('PBR') is None and fnguide_data.get('pbr'):
                        row['PBR'] = fmt(fnguide_data['pbr'])
                    if row.get('ROE(%)') is None and fnguide_data.get('roe'):
                        row['ROE(%)'] = fmt(fnguide_data['roe'])
                    if row.get('ROA(%)') is None and fnguide_data.get('roa'):
                        row['ROA(%)'] = fmt(fnguide_data['roa'])
                    if row.get('OpMargin(%)') is None and fnguide_data.get('op_margin'):
                        row['OpMargin(%)'] = fmt(fnguide_data['op_margin'])
                    if row.get('NetMargin(%)') is None and fnguide_data.get('net_margin'):
                        row['NetMargin(%)'] = fmt(fnguide_data['net_margin'])
                    if row.get('DebtRatio(%)') is None and fnguide_data.get('debt_ratio'):
                        row['DebtRatio(%)'] = fmt(fnguide_data['debt_ratio'])
                    if row.get('CurrentRatio') is None and fnguide_data.get('current_ratio'):
                        row['CurrentRatio'] = fmt(fnguide_data['current_ratio'])
                    if row.get('RevenueGrowth(%)') is None and fnguide_data.get('revenue_growth'):
                        row['RevenueGrowth(%)'] = fmt(fnguide_data['revenue_growth'])
                    if row.get('EPS') is None and fnguide_data.get('eps'):
                        row['EPS'] = fmt(fnguide_data['eps'], 0)
                    if row.get('BPS') is None and fnguide_data.get('bps'):
                        row['BPS'] = fmt(fnguide_data['bps'], 0)

            # ========================================
            # 3. pykrx (결측값 폴백)
            # ========================================
            if PYKRX_AVAILABLE:
                try:
                    today_str = datetime.now().strftime("%Y%m%d")
                    ohlcv = pykrx_stock.get_market_ohlcv_by_date(
                        (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
                        today_str, ticker
                    )
                    if not ohlcv.empty:
                        if row.get('Price') is None:
                            row['Price'] = fmt(ohlcv['종가'].iloc[-1], 0)
                        if row.get('Volume') is None:
                            row['Volume'] = int(ohlcv['거래량'].iloc[-1]) if ohlcv['거래량'].iloc[-1] else None
                except:
                    pass

            # ========================================
            # 4. ★ v3.0.2: FinanceDataReader 우선 (기술적 지표)
            # ========================================
            hist = pd.DataFrame()
            fdr_data = None
            
            if FDR_AVAILABLE:
                try:
                    fdr_data = fdr.DataReader(ticker, DATE_1Y_AGO)
                    if fdr_data is not None and not fdr_data.empty and len(fdr_data) > 20:
                        hist = fdr_data
                        if 'Close' not in hist.columns and '종가' in hist.columns:
                            hist = hist.rename(columns={'종가': 'Close', '시가': 'Open', '고가': 'High', '저가': 'Low', '거래량': 'Volume'})
                        log(f"    {ticker}: FDR 데이터 {len(hist)}일") if i == 0 else None
                except:
                    pass

            # ========================================
            # 5. yfinance (폴백)
            # ========================================
            info = {}
            
            if hist.empty:
                # v4.0: KOSPI(.KS) 또는 KOSDAQ(.KQ) 둘 다 시도
                ticker_variants = [f"{ticker}.KS", f"{ticker}.KQ"]
                t, hist, info = try_multiple_tickers(ticker_variants, max_retries=1)
            else:
                # FDR 성공 시에도 yfinance info는 가져옴 (재무 데이터용)
                yf_suffix = '.KQ' if market == 'KOSDAQ' else '.KS'
                try:
                    t = yf.Ticker(f"{ticker}{yf_suffix}")
                    info = t.info or {}
                except:
                    # 반대 시장도 시도
                    try:
                        alt_suffix = '.KS' if yf_suffix == '.KQ' else '.KQ'
                        t = yf.Ticker(f"{ticker}{alt_suffix}")
                        info = t.info or {}
                    except:
                        info = {}

            if not row.get('Price'):
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))
            if not row.get('Sector'):
                row['Sector'] = safe_get(info, 'sector')
            if row.get('MarketCap(억)') is None:
                cap = safe_get(info, 'marketCap')
                if cap:
                    row['MarketCap(억)'] = fmt(cap / 1e8, 0)
            
            if row.get('PER') is None:
                row['PER'] = fmt(safe_get(info, 'trailingPE'))
            if row.get('ForwardPE') is None:
                row['ForwardPE'] = fmt(safe_get(info, 'forwardPE'))
            if row.get('PBR') is None:
                row['PBR'] = fmt(safe_get(info, 'priceToBook'))
            
            if row.get('ROE(%)') is None and safe_get(info, 'returnOnEquity'):
                row['ROE(%)'] = fmt(safe_get(info, 'returnOnEquity') * 100)
            if row.get('ROA(%)') is None and safe_get(info, 'returnOnAssets'):
                row['ROA(%)'] = fmt(safe_get(info, 'returnOnAssets') * 100)
            if row.get('OpMargin(%)') is None and safe_get(info, 'operatingMargins'):
                row['OpMargin(%)'] = fmt(safe_get(info, 'operatingMargins') * 100)
            if row.get('GrossMargin(%)') is None and safe_get(info, 'grossMargins'):
                row['GrossMargin(%)'] = fmt(safe_get(info, 'grossMargins') * 100)
            
            if row.get('RevenueGrowth(%)') is None and safe_get(info, 'revenueGrowth'):
                row['RevenueGrowth(%)'] = fmt(safe_get(info, 'revenueGrowth') * 100)
            if row.get('EarningsGrowth(%)') is None and safe_get(info, 'earningsGrowth'):
                row['EarningsGrowth(%)'] = fmt(safe_get(info, 'earningsGrowth') * 100)

            if row.get('CurrentRatio') is None:
                row['CurrentRatio'] = fmt(safe_get(info, 'currentRatio'))
            if row.get('DebtRatio(%)') is None:
                row['DebtRatio(%)'] = fmt(safe_get(info, 'debtToEquity'))
            
            if row.get('DivYield(%)') is None and safe_get(info, 'dividendYield'):
                row['DivYield(%)'] = fmt(safe_get(info, 'dividendYield') * 100)
            
            # ★ v3.0: 기술적 지표 공통 함수 사용 (KR_Stocks는 MA60/120 포함)
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                row = add_technical_indicators(row, close, include_ma60_120=True)

            # 최종 폴백
            kr_field_mapping = {
                'PER': (('trailingPE', 'forwardPE'), 1, 2),
                'PBR': (('priceToBook',), 1, 2),
                'ROE(%)': (('returnOnEquity',), 100, 2),
                'ROA(%)': (('returnOnAssets',), 100, 2),
                'DivYield(%)': (('dividendYield',), 100, 2),
                'Beta': (('beta',), 1, 2),
                'MarketCap(억)': (('marketCap',), 1e-8, 0),
            }
            row = fill_missing_from_info(row, info, kr_field_mapping)

            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            # ★ v4.0: 데이터 검증 및 이상치 보정
            row = validate_and_clean_row(row, 'stock')
            row = calc_data_quality_score(row, REQUIRED_COLS_KR_STOCK, 'kr_stock')
            results.append(row)

        except Exception as e:
            results.append({'Code': ticker, 'Name': name, 'Market': market, 'Remark': str(e)[:30]})

        if (i + 1) % 10 == 0 or i == 0:
            elapsed = time.time() - start_time
            per_stock = elapsed / (i + 1) if i > 0 else 0
            remaining = per_stock * (len(all_tickers) - i - 1)
            log(f"  진행: {i+1}/{len(all_tickers)} ({(i+1)/len(all_tickers)*100:.0f}%) - 남은시간: {remaining/60:.1f}분")

        time.sleep(0.02)

    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# ★ v4.0 추가: 동적 종목 리스트 조회 함수
# ============================================================
def fetch_sp500_tickers():
    """위키피디아에서 S&P500 전체 종목 리스트 동적 조회"""
    try:
        log("  S&P500 종목 리스트 동적 조회 중 (Wikipedia)...")
        url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            table = soup.select_one('table.wikitable')
            if table:
                tickers = []
                for tr in table.select('tr')[1:]:
                    tds = tr.select('td')
                    if len(tds) >= 2:
                        ticker = tds[0].get_text().strip().replace('.', '-')
                        tickers.append(ticker)
                if len(tickers) > 400:
                    log(f"  ✅ S&P500 동적 조회: {len(tickers)}개")
                    return tickers
    except Exception as e:
        log(f"  ⚠️ S&P500 위키 조회 실패: {e}")

    # 2순위: FinanceDataReader
    if FDR_AVAILABLE:
        try:
            log("  FDR에서 S&P500 조회 시도...")
            sp500 = fdr.StockListing('S&P500')
            if sp500 is not None and not sp500.empty:
                tickers = sp500['Symbol'].tolist()
                if len(tickers) > 400:
                    log(f"  ✅ FDR S&P500: {len(tickers)}개")
                    return tickers
        except Exception as e:
            log(f"  ⚠️ FDR S&P500 실패: {e}")

    return []

def fetch_us_etf_tickers(max_count=300):
    """동적으로 미국 ETF 티커 목록 조회 (etfdb.com 또는 FDR)"""
    # 1순위: FinanceDataReader
    if FDR_AVAILABLE:
        try:
            log("  FDR에서 US ETF 조회 시도...")
            etf_list = fdr.StockListing('ETF/US')
            if etf_list is not None and not etf_list.empty:
                tickers = etf_list['Symbol'].head(max_count).tolist()
                if len(tickers) > 50:
                    log(f"  ✅ FDR US ETF: {len(tickers)}개")
                    return tickers
        except Exception as e:
            log(f"  ⚠️ FDR US ETF 실패: {e}")

    # 2순위: yfinance screener (주요 ETF)
    try:
        log("  yfinance에서 주요 US ETF 조회 시도...")
        # 카테고리별 주요 ETF 종합 (동적 확장 가능)
        major_etfs = []
        etf_categories = {
            'broad_market': ['SPY', 'IVV', 'VOO', 'VTI', 'QQQ', 'DIA', 'IWM', 'IWF', 'IWD', 'VUG', 'VTV',
                             'IJH', 'IJR', 'VB', 'VO', 'RSP', 'SPLG', 'SCHX', 'SCHB', 'MGK', 'VT', 'ACWI',
                             'ITOT', 'SPTM', 'SCHK', 'IWB', 'IWR', 'IWS', 'IWN', 'IWO', 'IWP'],
            'sector': ['XLK', 'XLF', 'XLV', 'XLE', 'XLI', 'XLY', 'XLP', 'XLU', 'XLB', 'XLRE', 'XLC',
                       'VGT', 'VFH', 'VHT', 'VDE', 'VIS', 'VCR', 'VDC', 'VPU', 'VAW', 'VNQ', 'VOX'],
            'thematic': ['ARKK', 'ARKW', 'ARKF', 'ARKG', 'SOXX', 'SMH', 'XBI', 'IBB', 'HACK', 'BOTZ',
                         'LIT', 'TAN', 'ICLN', 'PBW', 'QCLN', 'REMX', 'COPX', 'URA', 'KWEB', 'CIBR',
                         'ROBO', 'DRIV', 'PRNT', 'AIEQ', 'AIQ', 'IRBO', 'SNSR', 'FINX'],
            'dividend': ['VYM', 'SCHD', 'DVY', 'HDV', 'SPHD', 'SPYD', 'VIG', 'DGRO', 'NOBL',
                         'SDY', 'FVD', 'DGRW', 'DTD', 'RDIV', 'DLN', 'DHS', 'PEY', 'DIV'],
            'bond': ['BND', 'AGG', 'TLT', 'IEF', 'SHY', 'LQD', 'HYG', 'JNK', 'TIP', 'VCIT', 'VCSH',
                     'GOVT', 'BNDX', 'EMB', 'MUB', 'VTEB', 'BSV', 'BIV', 'BLV', 'SCHZ', 'FLOT', 'STIP'],
            'commodity': ['GLD', 'IAU', 'SLV', 'USO', 'UNG', 'DBC', 'PDBC', 'GLDM', 'SGOL', 'BAR', 'PPLT'],
            'leveraged': ['TQQQ', 'SQQQ', 'UPRO', 'SPXU', 'SOXL', 'SOXS', 'TNA', 'TZA', 'LABU', 'LABD'],
            'international': ['VEA', 'VWO', 'EFA', 'EEM', 'IEFA', 'IEMG', 'VXUS', 'IXUS', 'SCZ', 'ACWX',
                              'FXI', 'EWJ', 'EWG', 'EWU', 'EWY', 'EWT', 'EWZ', 'INDA', 'VGK', 'VPL'],
            'factor': ['MTUM', 'VLUE', 'QUAL', 'SIZE', 'USMV', 'EFAV', 'MOAT', 'COWZ', 'DSTL', 'GVAL'],
        }
        for cat, etfs in etf_categories.items():
            major_etfs.extend(etfs)
        # 중복 제거
        seen = set()
        unique = []
        for t in major_etfs:
            if t not in seen:
                seen.add(t)
                unique.append(t)
        log(f"  ✅ 확장된 US ETF 리스트: {len(unique)}개")
        return unique[:max_count]
    except Exception as e:
        log(f"  ⚠️ US ETF 리스트 생성 실패: {e}")

    return []

# ============================================================
# 미국 주식 수집
# ============================================================
# ★ v4.0: 하드코딩은 최후의 폴백으로만 사용
SP500_FALLBACK = [
    'AAPL', 'MSFT', 'AMZN', 'NVDA', 'GOOGL', 'META', 'TSLA', 'BRK-B', 'UNH', 'JNJ',
    'XOM', 'JPM', 'V', 'PG', 'MA', 'HD', 'CVX', 'MRK', 'ABBV', 'LLY',
    'PEP', 'KO', 'COST', 'AVGO', 'TMO', 'MCD', 'WMT', 'CSCO', 'ACN', 'ABT',
    'DHR', 'NEE', 'VZ', 'ADBE', 'CRM', 'NKE', 'PM', 'TXN', 'WFC', 'BMY',
    'RTX', 'CMCSA', 'ORCL', 'UPS', 'HON', 'QCOM', 'COP', 'T', 'LOW', 'MS',
    'INTC', 'UNP', 'ELV', 'BA', 'SPGI', 'CAT', 'IBM', 'GS', 'PLD', 'INTU',
    'DE', 'AMD', 'SBUX', 'AXP', 'AMAT', 'MDLZ', 'GE', 'BLK', 'GILD', 'ADI',
    'LMT', 'ISRG', 'TJX', 'SYK', 'CVS', 'BKNG', 'ADP', 'MMC', 'VRTX', 'REGN',
    'PGR', 'CB', 'NOW', 'CI', 'SCHW', 'ZTS', 'MO', 'TMUS', 'SO', 'DUK',
    'EOG', 'BDX', 'C', 'PNC', 'CL', 'TGT', 'ITW', 'SLB', 'AMT', 'USB',
]

def get_us_stocks():
    """미국 주식 데이터 - v4.0: S&P500 전체 동적 조회"""
    log("\n[2/6] 미국 주식 수집 중...")

    # ★ v4.0: 동적으로 S&P500 전체 리스트 조회
    tickers = fetch_sp500_tickers()

    # 폴백: 하드코딩 리스트
    if not tickers:
        log("  ⚠️ 동적 조회 실패, 하드코딩 폴백 사용")
        tickers = SP500_FALLBACK

    if TOP_N_US and len(tickers) > TOP_N_US:
        tickers = tickers[:TOP_N_US]
    results = []
    start_time = time.time()

    log(f"  대상: {len(tickers)}개")
    
    for i, ticker in enumerate(tickers):
        try:
            t, hist, info = fetch_yf_with_retry(ticker, max_retries=3, delay=0.5)

            if hist.empty and not info:
                continue

            row = {'Ticker': ticker}
            
            row['Name'] = safe_get(info, 'shortName', 'longName')
            row['Sector'] = safe_get(info, 'sector')
            row['Industry'] = safe_get(info, 'industry')
            row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))
            
            cap = safe_get(info, 'marketCap')
            if cap:
                row['MarketCap(B)'] = fmt(cap / 1e9, 1)
            
            # 밸류에이션
            row['PER'] = fmt(safe_get(info, 'trailingPE'))
            row['ForwardPE'] = fmt(safe_get(info, 'forwardPE'))
            row['PBR'] = fmt(safe_get(info, 'priceToBook'))
            row['PSR'] = fmt(safe_get(info, 'priceToSalesTrailing12Months'))
            row['PEG'] = fmt(safe_get(info, 'pegRatio'))
            row['EV/EBITDA'] = fmt(safe_get(info, 'enterpriseToEbitda'))
            
            # 수익성
            row['ROE(%)'] = fmt(safe_get(info, 'returnOnEquity', default=0) * 100) if safe_get(info, 'returnOnEquity') else None
            row['ROA(%)'] = fmt(safe_get(info, 'returnOnAssets', default=0) * 100) if safe_get(info, 'returnOnAssets') else None
            row['GrossMargin(%)'] = fmt(safe_get(info, 'grossMargins', default=0) * 100) if safe_get(info, 'grossMargins') else None
            row['OpMargin(%)'] = fmt(safe_get(info, 'operatingMargins', default=0) * 100) if safe_get(info, 'operatingMargins') else None
            row['NetMargin(%)'] = fmt(safe_get(info, 'profitMargins', default=0) * 100) if safe_get(info, 'profitMargins') else None
            
            # FCF Yield
            fcf = safe_get(info, 'freeCashflow')
            cap = safe_get(info, 'marketCap')
            if fcf and cap and cap > 0:
                row['FCFYield(%)'] = fmt(fcf / cap * 100)
            
            # 성장성
            row['RevenueGrowth(%)'] = fmt(safe_get(info, 'revenueGrowth', default=0) * 100) if safe_get(info, 'revenueGrowth') else None
            row['EarningsGrowth(%)'] = fmt(safe_get(info, 'earningsGrowth', default=0) * 100) if safe_get(info, 'earningsGrowth') else None
            
            # 안정성
            row['CurrentRatio'] = fmt(safe_get(info, 'currentRatio'))
            row['QuickRatio'] = fmt(safe_get(info, 'quickRatio'))
            row['DebtRatio(%)'] = fmt(safe_get(info, 'debtToEquity'))
            row['Debt/Equity'] = fmt(safe_get(info, 'debtToEquity'))
            
            # 배당
            row['DivYield(%)'] = fmt(safe_get(info, 'dividendYield', default=0) * 100) if safe_get(info, 'dividendYield') else None
            row['PayoutRatio(%)'] = fmt(safe_get(info, 'payoutRatio', default=0) * 100) if safe_get(info, 'payoutRatio') else None
            
            # 베타
            row['Beta'] = fmt(safe_get(info, 'beta'))
            
            # 기관 보유 비율
            inst = safe_get(info, 'heldPercentInstitutions')
            if inst:
                row['InstOwn(%)'] = fmt(inst * 100)
            
            # ★ v3.0 추가: EPS, BPS, Volume
            row['EPS'] = fmt(safe_get(info, 'trailingEps'))
            row['BPS'] = fmt(safe_get(info, 'bookValue'))
            row['Volume'] = safe_get(info, 'regularMarketVolume') or safe_get(info, 'averageVolume')
            
            # ★ v3.0: 기술적 지표 공통 함수 사용 (MA60/120 포함!)
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                row = add_technical_indicators(row, close, include_ma60_120=True)  # ★ MA60/120 추가
                
                # YTD 수익률
                try:
                    ytd_start = close[close.index >= f"{datetime.now().year}-01-01"]
                    if len(ytd_start) > 1:
                        row['ReturnYTD(%)'] = fmt((close.iloc[-1] / ytd_start.iloc[0] - 1) * 100)
                except:
                    pass

            # 최종 폴백
            us_field_mapping = {
                'PER': (('trailingPE', 'forwardPE'), 1, 2),
                'ForwardPE': (('forwardPE',), 1, 2),
                'PBR': (('priceToBook',), 1, 2),
                'ROE(%)': (('returnOnEquity',), 100, 2),
                'ROA(%)': (('returnOnAssets',), 100, 2),
                'DivYield(%)': (('dividendYield',), 100, 2),
                'Beta': (('beta',), 1, 2),
                'GrossMargin(%)': (('grossMargins',), 100, 2),
                'OpMargin(%)': (('operatingMargins',), 100, 2),
                'NetMargin(%)': (('profitMargins',), 100, 2),
            }
            row = fill_missing_from_info(row, info, us_field_mapping)

            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            if row.get('Price') is None:
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice', 'previousClose', 'open'))

            # ★ v4.0: 데이터 검증 및 이상치 보정
            row = validate_and_clean_row(row, 'stock')
            row = calc_data_quality_score(row, REQUIRED_COLS_US_STOCK, 'us_stock')
            results.append(row)

        except Exception as e:
            results.append({'Ticker': ticker, 'Remark': str(e)[:30]})

        if (i + 1) % 20 == 0 or i == 0:
            elapsed = time.time() - start_time
            per_stock = elapsed / (i + 1) if i > 0 else 0
            remaining = per_stock * (len(tickers) - i - 1)
            log(f"  진행: {i+1}/{len(tickers)} ({(i+1)/len(tickers)*100:.0f}%) - 남은시간: {remaining/60:.1f}분")

        time.sleep(0.15)

    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# ETF 공통 함수
# ============================================================
def get_etf_data(tickers, region=""):
    """ETF 데이터 수집 공통 함수 - v3.0 수정"""
    results = []

    for i, ticker in enumerate(tickers):
        try:
            t, hist, info = fetch_yf_with_retry(ticker, max_retries=3, delay=0.5)

            if hist.empty and not info:
                continue
            
            row = {'Ticker': ticker, 'Region': region}
            
            row['Name'] = safe_get(info, 'shortName', 'longName')
            row['Category'] = safe_get(info, 'category')
            row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))
            
            # ★ v3.0 수정: ExpenseRatio 하드코딩 우선
            yf_expense = safe_get(info, 'expenseRatio')
            if yf_expense:
                row['ExpenseRatio(%)'] = fmt(yf_expense * 100, 3)
            elif ticker in US_ETF_EXPENSE:  # ★ 하드코딩 폴백
                row['ExpenseRatio(%)'] = US_ETF_EXPENSE[ticker]
            else:
                row['ExpenseRatio(%)'] = None
            
            row['TotalAssets(B)'] = fmt(safe_get(info, 'totalAssets', default=0) / 1e9, 2) if safe_get(info, 'totalAssets') else None
            row['DivYield(%)'] = fmt(safe_get(info, 'yield', default=0) * 100) if safe_get(info, 'yield') else None
            
            # ★ v3.0: 기술적 지표 공통 함수 사용
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                row = add_technical_indicators(row, close, include_ma60_120=False)

            # 최종 폴백
            etf_field_mapping = {
                'DivYield(%)': (('yield', 'dividendYield'), 100, 2),
                'TotalAssets(B)': (('totalAssets',), 1e-9, 2),
            }
            row = fill_missing_from_info(row, info, etf_field_mapping)

            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            if row.get('Price') is None:
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice', 'previousClose', 'navPrice'))

            # ★ v4.0: 데이터 검증
            row = validate_and_clean_row(row, 'etf')
            row = calc_data_quality_score(row, REQUIRED_COLS_ETF, 'etf')
            results.append(row)

        except Exception as e:
            results.append({'Ticker': ticker, 'Remark': str(e)[:30]})

        if (i + 1) % 20 == 0 or i == 0:
            log(f"  진행: {i+1}/{len(tickers)}")

        time.sleep(0.15)

    return pd.DataFrame(results)

# ============================================================
# 한국 ETF
# ============================================================
def get_korea_etfs():
    """한국 ETF 데이터 - v4.0: FDR 우선, 검증 강화"""
    log("\n[3/6] 한국 ETF 수집 중...")
    
    # ★ v3.0.2: FinanceDataReader 최우선
    kr_etf_list = []
    if FDR_AVAILABLE:
        try:
            log("  FinanceDataReader에서 ETF 리스트 로드 중...")
            kr_etf_list = fetch_fdr_etf_list()
            if kr_etf_list:
                log(f"  FinanceDataReader: {len(kr_etf_list)}개 ETF")
        except Exception as e:
            log(f"  ⚠️ FinanceDataReader ETF 실패: {e}")
    
    # 2순위: KRX
    krx_data = {}
    if not kr_etf_list:
        krx_data = get_krx_etf_data()
        if krx_data:
            kr_etf_list = list(krx_data.keys())
            log(f"  KRX: {len(kr_etf_list)}개 ETF")
    
    # 3순위: pykrx
    if not kr_etf_list and PYKRX_AVAILABLE:
        try:
            log("  pykrx에서 ETF 리스트 로드 중...")
            kr_etf_list = pykrx_stock.get_etf_ticker_list()
            log(f"  pykrx: {len(kr_etf_list)}개 ETF")
        except Exception as e:
            log(f"  ⚠️ pykrx ETF 실패: {e}")
    
    # 4순위: 네이버
    if not kr_etf_list and NAVER_AVAILABLE is not False:
        naver_etfs = get_naver_etf_list(max_pages=10)
        kr_etf_list = [etf['code'] for etf in naver_etfs]
    
    # 하드코딩 폴백
    if not kr_etf_list and not krx_data:
        log("  ⚠️ ETF 리스트 로드 실패, 주요 ETF만 사용")
        kr_etf_list = list(KR_ETF_EXPENSE.keys())
    
    if TOP_N_KR_ETF and len(kr_etf_list) > TOP_N_KR_ETF:
        kr_etf_list = kr_etf_list[:TOP_N_KR_ETF]
    
    log(f"  대상: {len(kr_etf_list)}개")
    
    results = []
    start_time = time.time()
    
    for i, code in enumerate(kr_etf_list):
        try:
            code = str(code).zfill(6)  # 6자리로 패딩
            row = {'Code': code, 'Region': 'KR'}

            # 1. KRX 데이터 먼저 적용
            krx_info = krx_data.get(code, {}) if krx_data else {}
            if krx_info:
                row['Name'] = krx_info.get('name', '')
                if krx_info.get('price'):
                    row['Price'] = fmt(krx_info['price'], 0)
                if krx_info.get('nav'):
                    row['NAV'] = fmt(krx_info['nav'], 0)
                if krx_info.get('volume'):
                    row['Volume'] = int(krx_info['volume'])

            # 2. pykrx 보충
            if PYKRX_AVAILABLE and not row.get('Name'):
                try:
                    row['Name'] = pykrx_stock.get_etf_ticker_name(code)
                    today_str = datetime.now().strftime("%Y%m%d")
                    ohlcv = pykrx_stock.get_etf_ohlcv_by_date(
                        (datetime.now() - timedelta(days=7)).strftime("%Y%m%d"),
                        today_str, code
                    )
                    if not ohlcv.empty:
                        if not row.get('Price'):
                            row['Price'] = fmt(ohlcv['종가'].iloc[-1], 0)
                        if not row.get('Volume'):
                            row['Volume'] = int(ohlcv['거래량'].iloc[-1]) if ohlcv['거래량'].iloc[-1] else None
                except:
                    pass

            # ★ v3.0.2: FinanceDataReader 우선 (기술적 지표)
            hist = pd.DataFrame()
            if FDR_AVAILABLE:
                try:
                    fdr_data = fdr.DataReader(code, DATE_1Y_AGO)
                    if fdr_data is not None and not fdr_data.empty and len(fdr_data) > 20:
                        hist = fdr_data
                        if 'Close' not in hist.columns and '종가' in hist.columns:
                            hist = hist.rename(columns={'종가': 'Close', '시가': 'Open', '고가': 'High', '저가': 'Low', '거래량': 'Volume'})
                except:
                    pass

            # 3. yfinance (폴백)
            info = {}
            if hist.empty:
                ticker_variants = [f"{code}.KS", f"{code}.KQ"]
                t, hist, info = try_multiple_tickers(ticker_variants, max_retries=1)
            else:
                try:
                    t = yf.Ticker(f"{code}.KS")
                    info = t.info or {}
                except:
                    info = {}

            if not row.get('Name'):
                row['Name'] = safe_get(info, 'shortName', 'longName') or code
            if not row.get('Price'):
                row['Price'] = fmt(safe_get(info, 'regularMarketPrice'))

            # ★ v3.0 수정: Category 하드코딩 우선
            yf_category = safe_get(info, 'category')
            if yf_category:
                row['Category'] = yf_category
            elif code in KR_ETF_CATEGORY:  # ★ 하드코딩 폴백
                row['Category'] = KR_ETF_CATEGORY[code]
            else:
                row['Category'] = None
            
            # ★ v3.0 수정: ExpenseRatio 하드코딩 우선
            yf_expense = safe_get(info, 'expenseRatio')
            if yf_expense:
                row['ExpenseRatio(%)'] = fmt(yf_expense * 100, 3)
            elif code in KR_ETF_EXPENSE:  # ★ 하드코딩 폴백
                row['ExpenseRatio(%)'] = KR_ETF_EXPENSE[code]
            else:
                row['ExpenseRatio(%)'] = None
            
            # ★ v3.0 추가: DivYield 하드코딩
            yf_yield = safe_get(info, 'yield') or safe_get(info, 'dividendYield')
            if yf_yield:
                row['DivYield(%)'] = fmt(yf_yield * 100, 2)
            elif code in KR_ETF_DIVYIELD:  # ★ 하드코딩 폴백
                row['DivYield(%)'] = KR_ETF_DIVYIELD[code]
            else:
                row['DivYield(%)'] = None
            
            # ★ v3.0 추가: TotalAssets (v3.0.1: 단위 수정 1e12 → 1e9)
            total_assets = safe_get(info, 'totalAssets')
            if total_assets:
                row['TotalAssets(B)'] = fmt(total_assets / 1e9, 2)  # Billion 단위
            
            # ★ v3.0: 기술적 지표 공통 함수 사용
            if not hist.empty and len(hist) > 20:
                close = hist['Close']
                row = add_technical_indicators(row, close, include_ma60_120=False)
            else:
                if not info:
                    continue

            # 최종 폴백
            kr_etf_field_mapping = {
                'DivYield(%)': (('yield', 'dividendYield'), 100, 2),
            }
            row = fill_missing_from_info(row, info, kr_etf_field_mapping)

            if row.get('52wHigh') is None:
                val = safe_get(info, 'fiftyTwoWeekHigh')
                if val:
                    row['52wHigh'] = fmt(val)
            if row.get('52wLow') is None:
                val = safe_get(info, 'fiftyTwoWeekLow')
                if val:
                    row['52wLow'] = fmt(val)

            # ★ v4.0: 데이터 검증 및 이상치 보정
            row = validate_and_clean_row(row, 'etf')
            row = calc_data_quality_score(row, REQUIRED_COLS_ETF, 'etf')
            results.append(row)

        except Exception as e:
            pass

        if (i + 1) % 20 == 0 or i == 0:
            elapsed = time.time() - start_time
            per_etf = elapsed / (i + 1) if i > 0 else 0
            remaining = per_etf * (len(kr_etf_list) - i - 1)
            log(f"  진행: {i+1}/{len(kr_etf_list)} - 남은시간: {remaining/60:.1f}분")

        time.sleep(0.05)

    log(f"  ✅ 완료: {len(results)}개")
    return pd.DataFrame(results)

# ============================================================
# 미국 ETF - v4.0: 동적 조회
# ============================================================
# ★ v4.0: 하드코딩은 최후의 폴백으로만 사용
US_ETF_FALLBACK = [
    'SPY', 'IVV', 'VOO', 'VTI', 'QQQ', 'DIA', 'IWM', 'IWF', 'IWD', 'VUG',
    'VTV', 'IJH', 'IJR', 'VB', 'VO', 'RSP', 'SPLG', 'SCHX', 'SCHB', 'MGK',
    'XLK', 'XLF', 'XLV', 'XLE', 'XLI', 'XLY', 'XLP', 'XLU', 'XLB', 'XLRE',
    'VGT', 'VFH', 'VHT', 'VDE', 'VIS', 'VCR', 'VDC', 'VPU', 'VAW', 'VNQ',
    'ARKK', 'ARKW', 'ARKF', 'ARKG', 'SOXX', 'SMH', 'XBI', 'IBB', 'HACK', 'BOTZ',
    'LIT', 'TAN', 'ICLN', 'PBW', 'QCLN', 'REMX', 'COPX', 'URA',
    'VYM', 'SCHD', 'DVY', 'HDV', 'SPHD', 'SPYD', 'VIG', 'DGRO', 'NOBL',
    'BND', 'AGG', 'TLT', 'IEF', 'SHY', 'LQD', 'HYG', 'JNK', 'TIP',
    'GLD', 'IAU', 'SLV', 'USO', 'UNG', 'DBC', 'PDBC',
    'TQQQ', 'SQQQ', 'UPRO', 'SPXU', 'SOXL', 'SOXS',
    'VEA', 'VWO', 'EFA', 'EEM', 'IEFA', 'IEMG', 'VXUS', 'ACWI'
]

def get_us_etfs():
    """미국 ETF 데이터 - v4.0: 동적 조회"""
    log("\n[4/6] 미국 ETF 수집 중...")

    # ★ v4.0: 동적 ETF 리스트 조회
    tickers = fetch_us_etf_tickers(max_count=TOP_N_US_ETF or 300)

    # 폴백: 하드코딩 리스트
    if not tickers:
        log("  ⚠️ 동적 조회 실패, 하드코딩 폴백 사용")
        tickers = US_ETF_FALLBACK

    if TOP_N_US_ETF and len(tickers) > TOP_N_US_ETF:
        tickers = tickers[:TOP_N_US_ETF]

    log(f"  대상: {len(tickers)}개")
    df = get_etf_data(tickers, "US")
    log(f"  ✅ 완료: {len(df)}개")
    return df

# ============================================================
# ★★★ v4.0 추가: 거시경제 지표 수집 함수 ★★★
# ============================================================
def fetch_fred_series(series_id, api_key=None):
    """
    FRED(Federal Reserve Economic Data)에서 경제 지표 가져오기
    API key 없이도 웹 스크래핑으로 최신값 조회 가능
    """
    # 1순위: FRED 웹페이지 스크래핑 (API key 불필요)
    try:
        url = f"https://fred.stlouisfed.org/series/{series_id}"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            # FRED 페이지의 최신값 추출
            meta = soup.select_one('span.series-meta-observation-value')
            if meta:
                val_text = meta.get_text().strip().replace(',', '')
                try:
                    return float(val_text)
                except:
                    pass
            # 대안: meta tag에서
            for meta_tag in soup.select('meta[name="description"]'):
                content = meta_tag.get('content', '')
                match = re.search(r'([\d,.]+)\s*(?:percent|%|$)', content, re.IGNORECASE)
                if match:
                    try:
                        return float(match.group(1).replace(',', ''))
                    except:
                        pass
    except:
        pass

    # 2순위: FRED API (key가 있을 경우)
    if api_key:
        try:
            url = f"https://api.stlouisfed.org/fred/series/observations?series_id={series_id}&api_key={api_key}&file_type=json&sort_order=desc&limit=1"
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                obs = data.get('observations', [])
                if obs:
                    val = obs[0].get('value', '.')
                    if val != '.':
                        return float(val)
        except:
            pass

    return None

def get_macroeconomic_indicators():
    """
    v4.0: 주요 거시경제 지표 수집
    - Fed Funds Rate (기준금리)
    - CPI (소비자물가지수 / 인플레이션)
    - Core PCE (핵심 PCE)
    - 실업률
    - GDP 성장률
    - M2 통화량 증감률
    - 소비자신뢰지수
    - 비농업고용 변화
    - AAII 투자심리
    """
    results = []
    log("  거시경제 지표 수집 중...")

    # FRED 시리즈 ID 매핑
    fred_indicators = {
        'FEDFUNDS': ('Fed Funds Rate', '금리', '%', '미 연준 기준금리 (실효)'),
        'DFEDTARU': ('Fed Funds Target (상단)', '금리', '%', 'FOMC 목표금리 상단'),
        'DFEDTARL': ('Fed Funds Target (하단)', '금리', '%', 'FOMC 목표금리 하단'),
        'CPIAUCSL': ('CPI (도시소비자)', '인플레이션', 'index', '소비자물가지수'),
        'CPILFESL': ('Core CPI (식품/에너지 제외)', '인플레이션', 'index', '근원 소비자물가지수'),
        'PCEPI': ('PCE 물가지수', '인플레이션', 'index', '개인소비지출 물가지수'),
        'PCEPILFE': ('Core PCE (식품/에너지 제외)', '인플레이션', 'index', 'Fed 선호 인플레이션 지표'),
        'UNRATE': ('실업률', '고용', '%', '미국 실업률 (U-3)'),
        'PAYEMS': ('비농업고용', '고용', '천명', '비농업 고용자 수'),
        'ICSA': ('신규 실업수당 청구', '고용', '건', '주간 신규 실업수당 청구건수'),
        'GDP': ('GDP (명목)', '성장', 'B$', '미국 국내총생산'),
        'A191RL1Q225SBEA': ('실질 GDP 성장률 (QoQ)', '성장', '%', '분기별 실질 GDP 성장률 (연환산)'),
        'M2SL': ('M2 통화량', '유동성', 'B$', 'M2 통화 공급량'),
        'UMCSENT': ('미시간 소비자심리', '심리', 'index', '미시간대 소비자신뢰지수'),
        'CSCICP03USM665S': ('OECD 소비자신뢰', '심리', 'index', 'OECD 소비자신뢰지수 (미국)'),
        'T10YIE': ('10년 기대인플레이션', '인플레이션', '%', '10년 BEI (Breakeven Inflation)'),
        'T5YIE': ('5년 기대인플레이션', '인플레이션', '%', '5년 BEI'),
        'BAMLH0A0HYM2': ('하이일드 스프레드', '신용', '%', 'ICE BofA 하이일드 OAS'),
        'BAMLC0A0CM': ('투자등급 스프레드', '신용', '%', 'ICE BofA 투자등급 OAS'),
        'DTWEXBGS': ('달러 실효환율', '환율', 'index', '무역가중 달러 지수 (광의)'),
        'MORTGAGE30US': ('30년 모기지 금리', '금리', '%', '30년 고정 모기지 금리'),
        'HOUST': ('주택착공건수', '부동산', '천호', '신규 주택착공건수 (연환산)'),
    }

    collected = 0
    for series_id, (name, category, unit, description) in fred_indicators.items():
        try:
            val = fetch_fred_series(series_id)
            if val is not None:
                row = {
                    'Category': category,
                    'Ticker': f'FRED:{series_id}',
                    'Name': name,
                    'Price': fmt(val, 2 if unit == '%' else (1 if unit in ['index', 'B$'] else 0)),
                    'Unit': unit,
                    'Description': description,
                    'Source': 'FRED',
                }
                results.append(row)
                collected += 1
        except:
            pass
        time.sleep(0.3)  # FRED 부하 방지

    log(f"  FRED 지표: {collected}/{len(fred_indicators)}개 수집")

    # 추가: tradingeconomics에서 최신 경제 캘린더 주요 수치
    try:
        log("  Trading Economics에서 추가 지표 수집 중...")
        te_indicators = {
            'https://tradingeconomics.com/united-states/inflation-cpi': ('CPI YoY', '인플레이션', '%', 'CPI 전년동기비'),
            'https://tradingeconomics.com/united-states/core-inflation-rate': ('Core CPI YoY', '인플레이션', '%', '근원 CPI 전년동기비'),
            'https://tradingeconomics.com/united-states/gdp-growth-annual': ('GDP YoY 성장률', '성장', '%', 'GDP 전년동기비 성장률'),
        }

        for url, (name, category, unit, description) in te_indicators.items():
            try:
                resp = requests.get(url, headers=HEADERS, timeout=8)
                if resp.status_code == 200:
                    # tradingeconomics 최신값 추출
                    soup = BeautifulSoup(resp.text, 'lxml')
                    # 'Last' 또는 'Actual' 값 찾기
                    for elem in soup.select('#aspnetForm td'):
                        text = elem.get_text().strip()
                        try:
                            val = float(text)
                            if -20 <= val <= 30:  # 경제지표 합리적 범위
                                # 이미 같은 이름의 지표가 있는지 확인
                                exists = any(r['Name'] == name for r in results)
                                if not exists:
                                    results.append({
                                        'Category': category,
                                        'Ticker': f'TE:{name}',
                                        'Name': name,
                                        'Price': fmt(val, 2),
                                        'Unit': unit,
                                        'Description': description,
                                        'Source': 'TradingEconomics',
                                    })
                                break
                        except:
                            pass
                time.sleep(0.5)
            except:
                pass
    except:
        pass

    log(f"  ✅ 거시경제 지표 총 {len(results)}개 수집")
    return results

# ============================================================
# 시장 지표
# ============================================================
def get_fear_greed_index():
    """CNN Fear & Greed Index 스크래핑"""
    try:
        url = "https://production.dataviz.cnn.io/index/fearandgreed/graphdata"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'application/json',
        }
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if 'fear_and_greed' in data:
                return {
                    'value': fmt(data['fear_and_greed'].get('score'), 1),
                    'rating': data['fear_and_greed'].get('rating', ''),
                    'previous_close': fmt(data['fear_and_greed'].get('previous_close'), 1),
                }
    except:
        pass
    return None

def get_cape_ratio():
    """S&P 500 Shiller CAPE Ratio from multpl.com"""
    try:
        url = "https://www.multpl.com/shiller-pe"
        resp = requests.get(url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            current = soup.select_one('#current')
            if current:
                val_text = current.get_text().strip().replace(',', '')
                match = re.search(r'([\d.]+)', val_text)
                if match:
                    return fmt(float(match.group(1)), 2)
    except:
        pass
    return None

def get_sp500_forward_pe():
    """S&P 500 Forward PE from SPY ETF"""
    try:
        for ticker in ['SPY', 'IVV', 'VOO']:
            t = yf.Ticker(ticker)
            info = t.info
            pe = safe_get(info, 'trailingPE')
            if pe and 10 < pe < 50:
                return fmt(pe, 2)
    except:
        pass
    return None

def get_ism_pmi():
    """ISM 제조업 PMI from tradingeconomics"""
    try:
        url = "https://tradingeconomics.com/united-states/business-confidence"
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_LONG)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, 'lxml')
            for elem in soup.select('#aspnetForm td, #aspnetForm span'):
                text = elem.get_text().strip()
                try:
                    val = float(text)
                    if 40 <= val <= 65:
                        return fmt(val, 1)
                except:
                    pass
    except:
        pass
    
    try:
        url = "https://www.investing.com/economic-calendar/ism-manufacturing-pmi-173"
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_LONG)
        if resp.status_code == 200:
            match = re.search(r'Actual.*?(\d+\.?\d*)', resp.text)
            if match:
                val = float(match.group(1))
                if 40 <= val <= 65:
                    return fmt(val, 1)
    except:
        pass
    
    return None

# ★★★ v4.1 추가: 한국 국고채 금리 + 한국은행 기준금리 + JGB 수집 ★★★
# ============================================================
def get_korea_bond_yields():
    """
    한국 국고채 수익률 + 한국은행 기준금리 수집
    v4.1: 투자 판단 시스템의 한국 금융 NIM 판단에 필요
    
    소스 우선순위:
    1. 네이버 금융 국고채 페이지
    2. FRED 한국 금리 시리즈
    3. FinanceDataReader
    """
    results = {}
    log("  한국 채권/금리 수집 중...")
    
    # ── 네이버 금융 채권 시세 (1순위) ──
    naver_bonds = {
        'IRR_GOVT01Y': ('kr_1y', '국고채 1년'),
        'IRR_GOVT03Y': ('kr_3y', '국고채 3년'),
        'IRR_GOVT10Y': ('kr_10y', '국고채 10년'),
        'IRR_CD91':    ('kr_cd91', 'CD 91일'),
    }
    
    for mkt_code, (key, label) in naver_bonds.items():
        try:
            url = f"https://finance.naver.com/marketindex/interestDailyQuote.naver?marketindexCd={mkt_code}"
            resp = requests.get(url, headers=HEADERS, timeout=8)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'lxml')
                tds = soup.select('table.tbl_exchange td')
                if tds:
                    val_text = tds[0].get_text().strip().replace(',', '')
                    try:
                        val = float(val_text)
                        if 0.5 < val < 10:
                            results[key] = val
                    except:
                        pass
            time.sleep(0.3)
        except:
            pass
    
    # ── FRED 한국 금리 (네이버 실패 시 폴백) ──
    fred_kr_map = {
        'IRLTLT01KRM156N': 'kr_10y',  # Korea Long-Term Rate (~10Y)
        'IRSTCI01KRM156N': 'bok_rate', # Korea Short-Term Interest Rate (기준금리 proxy)
    }
    for series_id, key in fred_kr_map.items():
        if not results.get(key):
            try:
                val = fetch_fred_series(series_id)
                if val and 0 < val < 10:
                    results[key] = val
            except:
                pass
            time.sleep(0.3)
    
    # ── FinanceDataReader 한국 국고채 (추가 폴백) ──
    if FDR_AVAILABLE:
        fdr_bond_map = {
            'KR3YT=RR': 'kr_3y',
            'KR10YT=RR': 'kr_10y',
            'KR1YT=RR': 'kr_1y',
        }
        end_date = datetime.now().strftime('%Y-%m-%d')
        start_date = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')
        
        for fdr_ticker, key in fdr_bond_map.items():
            if not results.get(key):
                try:
                    df = fdr.DataReader(fdr_ticker, start_date, end_date)
                    if df is not None and not df.empty:
                        val = float(df['Close'].iloc[-1])
                        if 0.5 < val < 10:
                            results[key] = val
                except:
                    pass
    
    # ── 한국은행 기준금리 추정 (직접 수집 실패 시) ──
    if not results.get('bok_rate') and results.get('kr_1y'):
        # 국고채 1Y와 기준금리 차이는 보통 0.1~0.3%p
        results['bok_rate'] = round(results['kr_1y'] - 0.15, 2)
    
    collected = len(results)
    log(f"  ✅ 한국 채권/금리: {collected}개 수집"
        f" (기준금리:{results.get('bok_rate','N/A')}"
        f", 국고3Y:{results.get('kr_3y','N/A')}"
        f", 국고10Y:{results.get('kr_10y','N/A')}"
        f", CD91:{results.get('kr_cd91','N/A')})")
    return results


def get_jgb_yield():
    """
    일본 JGB 10년 수익률 (JGB 환류 리스크 판단용)
    v4.1: JGB 10Y > 1.5%이면 미국 장기채 매도 압력 발생
    """
    # 1순위: FRED Japan Long-Term Rate
    try:
        val = fetch_fred_series('IRLTLT01JPM156N')
        if val and 0 < val < 5:
            log(f"  ✅ JGB 10Y: {val}% (FRED)")
            return val
    except:
        pass
    
    # 2순위: yfinance (메인 루프에서 수집 실패 시)
    try:
        t = yf.Ticker('^JGBS')
        hist = t.history(period='5d')
        if not hist.empty:
            val = float(hist['Close'].iloc[-1])
            if 0 < val < 5:
                log(f"  ✅ JGB 10Y: {val}% (yfinance ^JGBS)")
                return val
    except:
        pass
    
    log("  ⚠️ JGB 10Y 수집 실패")
    return None


def get_korea_market_indicators():
    """한국 시장 지표 (pykrx + 네이버 대안)"""
    indicators = {}

    if PYKRX_AVAILABLE:
        try:
            today_str = datetime.now().strftime("%Y%m%d")

            try:
                from_date = (datetime.now() - timedelta(days=7)).strftime("%Y%m%d")
                kospi_fund = pykrx_stock.get_index_fundamental(from_date, today_str, "1001")
                if not kospi_fund.empty:
                    latest = kospi_fund.iloc[-1]
                    if 'PER' in kospi_fund.columns:
                        indicators['kospi_per'] = fmt(latest['PER'], 2)
                    if 'PBR' in kospi_fund.columns:
                        indicators['kospi_pbr'] = fmt(latest['PBR'], 2)
                    if 'DIV' in kospi_fund.columns:
                        indicators['kospi_div'] = fmt(latest['DIV'], 2)
            except Exception as e:
                log(f"  ⚠️ pykrx 펀더멘털 실패: {e}")

            try:
                from_date = (datetime.now() - timedelta(days=30)).strftime("%Y%m%d")
                investor = pykrx_stock.get_market_trading_value_by_date(from_date, today_str, "KOSPI")
                if not investor.empty and '외국인' in investor.columns:
                    recent_5 = investor['외국인'].tail(5).sum()
                    indicators['foreign_net_buy'] = fmt(recent_5 / 1e8, 0)
                    recent_20 = investor['외국인'].tail(20).sum()
                    indicators['foreign_net_buy_20d'] = fmt(recent_20 / 1e8, 0)
                    
                if not investor.empty:
                    if '개인' in investor.columns:
                        indicators['individual_net_buy'] = fmt(investor['개인'].tail(5).sum() / 1e8, 0)
                    if '기관합계' in investor.columns:
                        indicators['institution_net_buy'] = fmt(investor['기관합계'].tail(5).sum() / 1e8, 0)
            except Exception as e:
                log(f"  ⚠️ pykrx 투자자 실패: {e}")

        except Exception as e:
            log(f"  ⚠️ pykrx 전체 실패: {e}")

    if not indicators.get('kospi_per') and NAVER_AVAILABLE is not False:
        try:
            url = "https://finance.naver.com/sise/sise_index.naver?code=KOSPI"
            resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT_SHORT)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, 'lxml')
                for td in soup.select('td'):
                    text = td.get_text().strip()
                    if 'PER' in text or 'per' in text.lower():
                        parent = td.find_parent('tr')
                        if parent:
                            tds = parent.select('td')
                            for t in tds:
                                val_text = t.get_text().strip().replace(',', '')
                                try:
                                    val = float(val_text)
                                    if 5 < val < 50:
                                        indicators['kospi_per'] = fmt(val, 2)
                                        break
                                except:
                                    pass
                log("  ✅ 네이버에서 코스피 지표 로드")
        except Exception as e:
            log(f"  ⚠️ 네이버 코스피 지표 실패: {e}")

    return indicators

def get_market_indicators():
    """글로벌 시장 지표 - v4.1: 한국 국고채·JGB·스프레드 추가"""
    log("\n[5/6] 시장 지표 수집 중...")

    indicators = {
        '^GSPC': ('S&P 500', '지수'),
        '^DJI': ('다우존스', '지수'),
        '^IXIC': ('나스닥 종합', '지수'),
        '^VIX': ('VIX 변동성', '지수'),
        '^KS11': ('코스피', '지수'),
        '^KQ11': ('코스닥', '지수'),
        '^N225': ('니케이 225', '지수'),
        '^HSI': ('항셍', '지수'),
        '^STOXX50E': ('유로스톡스50', '지수'),
        '^FTSE': ('FTSE 100', '지수'),

        'USDKRW=X': ('USD/KRW', '환율'),
        'EURUSD=X': ('EUR/USD', '환율'),
        'USDJPY=X': ('USD/JPY', '환율'),
        'DX-Y.NYB': ('달러 인덱스', '환율'),
        'USDCNY=X': ('USD/CNY', '환율'),

        'GC=F': ('금 선물', '원자재'),
        'SI=F': ('은 선물', '원자재'),
        'CL=F': ('WTI 원유', '원자재'),
        'BZ=F': ('브렌트유', '원자재'),
        'NG=F': ('천연가스', '원자재'),
        'HG=F': ('구리 선물', '원자재'),

        '^IRX': ('미국채 3개월', '채권'),
        '^FVX': ('미국채 5년', '채권'),
        '^TNX': ('미국채 10년', '채권'),
        '^TYX': ('미국채 30년', '채권'),

        # ★ v4.1: 투자 판단 시스템 핵심 채권 지표
        '2YY=F': ('미국채 2년', '채권'),         # 10Y-2Y 스프레드 계산용

        'BTC-USD': ('비트코인', '암호화폐'),
        'ETH-USD': ('이더리움', '암호화폐'),

        'HYG': ('하이일드 채권 ETF', '신용'),
        'LQD': ('투자등급 채권 ETF', '신용'),
        'TLT': ('장기국채 ETF', '채권'),
        'SHY': ('단기국채 ETF', '채권'),

        '^CPCE': ('Put/Call Ratio', '심리'),
    }

    results = []

    log("  Yahoo Finance 지표 수집 중...")
    for ticker, (name, category) in indicators.items():
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="1y")

            if hist.empty:
                continue

            close = hist['Close']
            current_price = close.iloc[-1]

            decimals = 4 if category == '환율' else (3 if category == '채권' else 2)

            row = {
                'Category': category,
                'Ticker': ticker,
                'Name': name,
                'Price': fmt(current_price, decimals),
            }

            if len(close) >= 2:
                row['Change(%)'] = fmt((close.iloc[-1] / close.iloc[-2] - 1) * 100)
            if len(close) >= 6:
                row['Return1W(%)'] = fmt((close.iloc[-1] / close.iloc[-6] - 1) * 100)
            if len(close) >= 22:
                row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)
            if len(close) >= 66:
                row['Return3M(%)'] = fmt((close.iloc[-1] / close.iloc[-66] - 1) * 100)
            if len(close) >= 245:  # ★ v3.0: 252 → 245
                row['Return1Y(%)'] = fmt((close.iloc[-1] / close.iloc[-min(len(close), 252)] - 1) * 100)  # ★ v3.0.1 버그 수정
                year_data = close.tail(min(len(close), 252))
                row['52wHigh'] = fmt(year_data.max(), decimals)
                row['52wLow'] = fmt(year_data.min(), decimals)
                row['From52wHigh(%)'] = fmt((current_price / year_data.max() - 1) * 100)
                row['From52wLow(%)'] = fmt((current_price / year_data.min() - 1) * 100)

            results.append(row)
        except:
            pass

        time.sleep(0.05)

    log("  계산 지표 수집 중...")

    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        irx = next((r for r in results if r['Ticker'] == '^IRX'), None)
        if tnx and irx and tnx.get('Price') and irx.get('Price'):
            spread_10y_3m = float(tnx['Price']) - float(irx['Price'])
            results.append({
                'Category': '스프레드',
                'Ticker': '10Y-3M',
                'Name': '10년-3개월 스프레드',
                'Price': fmt(spread_10y_3m, 3),
                'Signal': '역전' if spread_10y_3m < 0 else '정상',
            })
    except:
        pass

    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        fvx = next((r for r in results if r['Ticker'] == '^FVX'), None)
        if tnx and fvx and tnx.get('Price') and fvx.get('Price'):
            spread_10y_5y = float(tnx['Price']) - float(fvx['Price'])
            results.append({
                'Category': '스프레드',
                'Ticker': '10Y-5Y',
                'Name': '10년-5년 스프레드',
                'Price': fmt(spread_10y_5y, 3),
            })
    except:
        pass

    try:
        hyg = next((r for r in results if r['Ticker'] == 'HYG'), None)
        lqd = next((r for r in results if r['Ticker'] == 'LQD'), None)
        if hyg and lqd and hyg.get('Return1M(%)') and lqd.get('Return1M(%)'):
            hy_spread_proxy = float(lqd['Return1M(%)']) - float(hyg['Return1M(%)'])
            results.append({
                'Category': '스프레드',
                'Ticker': 'HY-IG',
                'Name': '하이일드 스프레드 (proxy)',
                'Price': fmt(hy_spread_proxy, 2),
                'Description': 'LQD-HYG 1개월 수익률 차이',
            })
    except:
        pass

    log("  심리 지표 수집 중...")

    fg_data = get_fear_greed_index()
    if fg_data:
        results.append({
            'Category': '심리',
            'Ticker': 'F&G',
            'Name': 'Fear & Greed Index',
            'Price': fg_data.get('value'),
            'Signal': fg_data.get('rating', ''),
            'Previous': fg_data.get('previous_close'),
        })

    cape = get_cape_ratio()
    if cape:
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'CAPE',
            'Name': 'Shiller CAPE Ratio',
            'Price': cape,
            'Signal': '고평가' if float(cape) > 30 else ('저평가' if float(cape) < 15 else '보통'),
        })

    forward_pe = get_sp500_forward_pe()
    if forward_pe:
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'SPY-PE',
            'Name': 'S&P500 PE (trailing)',
            'Price': forward_pe,
        })

    ism_pmi = get_ism_pmi()
    if ism_pmi:
        results.append({
            'Category': '경기',
            'Ticker': 'ISM-PMI',
            'Name': 'ISM 제조업 PMI',
            'Price': ism_pmi,
            'Signal': '확장' if float(ism_pmi) > 50 else '수축',
        })

    log("  한국 시장 지표 수집 중...")

    kr_indicators = get_korea_market_indicators()

    if kr_indicators.get('kospi_per'):
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'KOSPI-PER',
            'Name': '코스피 PER',
            'Price': kr_indicators['kospi_per'],
        })

    if kr_indicators.get('kospi_pbr'):
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'KOSPI-PBR',
            'Name': '코스피 PBR',
            'Price': kr_indicators['kospi_pbr'],
        })

    if kr_indicators.get('kospi_div'):
        results.append({
            'Category': '밸류에이션',
            'Ticker': 'KOSPI-DIV',
            'Name': '코스피 배당수익률',
            'Price': kr_indicators['kospi_div'],
        })

    if kr_indicators.get('foreign_net_buy') is not None:
        val = kr_indicators['foreign_net_buy']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-외국인',
            'Name': '외국인 순매수 (5일, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    if kr_indicators.get('foreign_net_buy_20d') is not None:
        val = kr_indicators['foreign_net_buy_20d']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-외국인-20D',
            'Name': '외국인 순매수 (20일 누적, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    if kr_indicators.get('individual_net_buy') is not None:
        val = kr_indicators['individual_net_buy']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-개인',
            'Name': '개인 순매수 (5일, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    if kr_indicators.get('institution_net_buy') is not None:
        val = kr_indicators['institution_net_buy']
        results.append({
            'Category': '수급',
            'Ticker': 'KR-기관',
            'Name': '기관 순매수 (5일, 억원)',
            'Price': val,
            'Signal': '매수' if float(val) > 0 else '매도',
        })

    log("  추가 ETF 지표 수집 중...")

    # ★★★ v4.1: 한국 국고채 금리 + 한국은행 기준금리 ★★★
    log("  한국 채권/금리 수집 중...")
    kr_bond_data = get_korea_bond_yields()
    
    if kr_bond_data.get('bok_rate') is not None:
        results.append({
            'Category': '금리',
            'Ticker': 'BOK-RATE',
            'Name': '한국은행 기준금리',
            'Price': fmt(kr_bond_data['bok_rate'], 2),
            'Unit': '%',
            'Description': '한국은행 기준금리. 한국 은행 NIM의 핵심 결정 변수',
        })
    
    if kr_bond_data.get('kr_1y') is not None:
        results.append({
            'Category': '채권',
            'Ticker': 'KR-1Y',
            'Name': '한국 국고채 1년',
            'Price': fmt(kr_bond_data['kr_1y'], 3),
            'Unit': '%',
        })
    
    if kr_bond_data.get('kr_3y') is not None:
        results.append({
            'Category': '채권',
            'Ticker': 'KR-3Y',
            'Name': '한국 국고채 3년',
            'Price': fmt(kr_bond_data['kr_3y'], 3),
            'Unit': '%',
            'Description': '한국 국고채 3년물. 한국 금리 환경의 핵심 지표',
        })
    
    if kr_bond_data.get('kr_10y') is not None:
        results.append({
            'Category': '채권',
            'Ticker': 'KR-10Y',
            'Name': '한국 국고채 10년',
            'Price': fmt(kr_bond_data['kr_10y'], 3),
            'Unit': '%',
        })
    
    if kr_bond_data.get('kr_cd91') is not None:
        results.append({
            'Category': '금리',
            'Ticker': 'KR-CD91',
            'Name': 'CD 91일 금리',
            'Price': fmt(kr_bond_data['kr_cd91'], 3),
            'Unit': '%',
            'Description': 'CD 91일 금리. 은행 단기 조달비용 proxy',
        })
    
    # 한국 수익률 곡선: 국고 10Y - 1Y 스프레드
    if kr_bond_data.get('kr_10y') and kr_bond_data.get('kr_1y'):
        kr_spread = kr_bond_data['kr_10y'] - kr_bond_data['kr_1y']
        results.append({
            'Category': '스프레드',
            'Ticker': 'KR-10Y-1Y',
            'Name': '한국 국고 10Y-1Y 스프레드',
            'Price': fmt(kr_spread, 3),
            'Signal': '역전' if kr_spread < 0 else '정상',
            'Description': '한국 수익률 곡선 기울기',
        })
    
    # ★★★ v4.1: JGB 10년 수익률 ★★★
    log("  일본 JGB 수집 중...")
    jgb_val = get_jgb_yield()
    if jgb_val is not None:
        results.append({
            'Category': '채권',
            'Ticker': 'JGB-10Y',
            'Name': '일본 JGB 10년',
            'Price': fmt(jgb_val, 3),
            'Unit': '%',
            'Description': '일본 국채 10년물. >1.5%이면 미국 장기채 환류 리스크',
            'Signal': '환류위험' if jgb_val > 1.5 else '안정',
        })
    
    # ★★★ v4.1: 미국 10Y-2Y 스프레드 (핵심 수익률 곡선 지표) ★★★
    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        us2y = next((r for r in results if r['Ticker'] == '2YY=F'), None)
        
        us2y_val = None
        if us2y and us2y.get('Price'):
            us2y_val = float(us2y['Price'])
        
        # 2Y 수집 실패 시 FRED 대안 (DGS2)
        if us2y_val is None:
            try:
                fred_2y = fetch_fred_series('DGS2')
                if fred_2y and 0 < fred_2y < 8:
                    us2y_val = fred_2y
                    results.append({
                        'Category': '채권',
                        'Ticker': 'US-2Y',
                        'Name': '미국채 2년 (FRED)',
                        'Price': fmt(fred_2y, 3),
                        'Unit': '%',
                        'Source': 'FRED',
                    })
            except:
                pass
        
        if tnx and tnx.get('Price') and us2y_val:
            spread_10y_2y = float(tnx['Price']) - us2y_val
            results.append({
                'Category': '스프레드',
                'Ticker': '10Y-2Y',
                'Name': '미국 10Y-2Y 스프레드',
                'Price': fmt(spread_10y_2y, 3),
                'Signal': '역전' if spread_10y_2y < 0 else ('정상화' if spread_10y_2y > 0 else '플랫'),
                'Description': '핵심 수익률 곡선 지표. 역전→정상화 전환 시 침체 경고',
            })
    except:
        pass
    
    # ★★★ v4.1: 한미 금리차 (원화 환율 방향 판단용) ★★★
    try:
        kr_3y_val = kr_bond_data.get('kr_3y')
        if us2y_val and kr_3y_val:
            rate_diff = us2y_val - kr_3y_val
            results.append({
                'Category': '스프레드',
                'Ticker': 'US2Y-KR3Y',
                'Name': '한미 금리차 (US2Y-KR3Y)',
                'Price': fmt(rate_diff, 3),
                'Signal': '원화약세 압력' if rate_diff > 1.0 else ('원화강세' if rate_diff < -0.5 else '보통'),
                'Description': '한미 금리차. >1%p이면 원화약세/외국인 이탈 압력',
            })
    except:
        pass

    additional_etfs = {
        'IEF': ('미국채 7-10년 ETF', '채권'),
        'TIP': ('물가연동채 ETF', '채권'),
        'EMB': ('신흥국 채권 ETF', '신용'),
        'GLD': ('금 ETF', '원자재'),
        'USO': ('원유 ETF', '원자재'),
        'VXX': ('VIX 선물 ETF', '변동성'),
        'SVXY': ('VIX 인버스 ETF', '변동성'),
    }

    for ticker, (name, category) in additional_etfs.items():
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="3mo")

            if hist.empty:
                continue

            close = hist['Close']
            current_price = close.iloc[-1]

            row = {
                'Category': category,
                'Ticker': ticker,
                'Name': name,
                'Price': fmt(current_price, 2),
            }

            if len(close) >= 22:
                row['Return1M(%)'] = fmt((close.iloc[-1] / close.iloc[-22] - 1) * 100)

            results.append(row)
        except:
            pass

        time.sleep(0.05)

    try:
        tnx = next((r for r in results if r['Ticker'] == '^TNX'), None)
        tip = next((r for r in results if r['Ticker'] == 'TIP'), None)
        if tnx and tip:
            tip_return = tip.get('Return1M(%)')
            if tip_return:
                real_rate = float(tnx['Price']) - (float(tip_return) * 12 / 100)
                results.append({
                    'Category': '금리',
                    'Ticker': 'REAL-RATE',
                    'Name': '실질금리 추정',
                    'Price': fmt(real_rate, 3),
                    'Description': '10Y - TIP implied inflation',
                })
    except:
        pass

    # ★★★ v4.0: 거시경제 지표 추가 ★★★
    log("\n[6/6] 거시경제 지표 수집 중...")
    macro_results = get_macroeconomic_indicators()
    results.extend(macro_results)

    log(f"  ✅ 시장+거시경제 지표 총 {len(results)}개 완료")
    return pd.DataFrame(results)

# ============================================================
# Excel 저장
# ============================================================
def save_to_excel(data_dict, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    log("\n엑셀 저장 중...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for sheet_name, df in data_dict.items():
        if df is None or df.empty:
            continue
        
        ws = wb.create_sheet(title=sheet_name[:31])
        
        for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ws.freeze_panes = 'A2'
        log(f"  ✅ {sheet_name}: {len(df)}행")
    
    wb.save(filename)
    log(f"\n💾 Excel 저장: {filename}")

# ============================================================
# JSON 저장
# ============================================================
def save_to_json(data_dict, filename):
    log("\nJSON 저장 중...")
    
    output = {
        'metadata': {
            'generated': datetime.now().isoformat(),
            'date': TODAY,
            'version': 'v4.1-bond-enhanced'  # ★ v4.0 업데이트
        },
        'data': {}
    }
    
    for sheet_name, df in data_dict.items():
        if df is None or df.empty:
            continue
        
        records = df.replace({np.nan: None}).to_dict(orient='records')
        output['data'][sheet_name] = records
        log(f"  ✅ {sheet_name}: {len(records)}개")
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    log(f"\n💾 JSON 저장: {filename}")
    
    size_mb = os.path.getsize(filename) / (1024 * 1024)
    log(f"   파일 크기: {size_mb:.2f} MB")

# ============================================================
# 메인
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='글로벌 주식/ETF 스크리너 (GitHub Actions) v4.1')
    parser.add_argument('--json-only', action='store_true', help='JSON만 출력')
    parser.add_argument('--output-dir', type=str, default='.', help='출력 디렉토리')
    parser.add_argument('--kr-stocks', type=int, default=500, help='한국 주식 수 (기본 500)')
    parser.add_argument('--us-stocks', type=int, default=500, help='미국 주식 수 (기본 500)')
    parser.add_argument('--kr-etfs', type=int, default=300, help='한국 ETF 수 (기본 300)')
    parser.add_argument('--us-etfs', type=int, default=300, help='미국 ETF 수 (기본 300)')
    args = parser.parse_args()

    global TOP_N_KR, TOP_N_US, TOP_N_KR_ETF, TOP_N_US_ETF
    TOP_N_KR = args.kr_stocks
    TOP_N_US = args.us_stocks
    TOP_N_KR_ETF = args.kr_etfs
    TOP_N_US_ETF = args.us_etfs

    log("=" * 60)
    log("글로벌 주식/ETF 스크리닝 - GitHub Actions v4.1")
    log("★ v4.1: 한국 국고채·JGB·US2Y·한미금리차 추가")
    log(f"실행: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"한국 주식: 최대 {TOP_N_KR}개 (KOSPI+KOSDAQ) | 미국 주식: 최대 {TOP_N_US}개 (S&P500 동적)")
    log(f"한국 ETF: 최대 {TOP_N_KR_ETF}개 | 미국 ETF: 최대 {TOP_N_US_ETF}개 (동적)")
    log("=" * 60)
    
    start = time.time()
    
    data = {
        'KR_Stocks': get_korea_stocks(),
        'US_Stocks': get_us_stocks(),
        'KR_ETF': get_korea_etfs(),
        'US_ETF': get_us_etfs(),
        'Market_Indicators': get_market_indicators()
    }
    
    # ========== v4.5: Phase B/C 통합 ==========
    # Phase B: 재무 이벤트 (DART 공시)
    if PHASE_B_AVAILABLE:
        try:
            log("\n🔄 Phase B: DART 재무 이벤트 수집 중...")
            data = integrate_phase_b_standalone(data)
            log(f"✅ Phase B: {len(data.get('FinancialEvents', []))}건 이벤트")
        except Exception as e:
            log(f"⚠️  Phase B 실패 (계속 진행): {e}")
            data.setdefault('FinancialEvents', [])
            data.setdefault('EarningsRevisions', {})
    else:
        data['FinancialEvents'] = []
        data['EarningsRevisions'] = {}
    
    # Phase C: 수급 시그널 (KRX pykrx)
    if PHASE_C_AVAILABLE:
        try:
            log("\n🔄 Phase C: KRX 수급 시그널 수집 중...")
            data = integrate_phase_c_standalone(data)
            log(f"✅ Phase C: {len(data.get('FlowSignals', {}))}종목 시그널")
        except Exception as e:
            log(f"⚠️  Phase C 실패 (계속 진행): {e}")
            data.setdefault('FlowSignals', {})
    else:
        data['FlowSignals'] = {}
    # ========== v4.5 통합 끝 ==========
    
    output_dir = args.output_dir
    os.makedirs(output_dir, exist_ok=True)
    
    json_file = os.path.join(output_dir, 'data.json')
    save_to_json(data, json_file)
    
    if not args.json_only:
        excel_file = os.path.join(output_dir, f'global_screening_{TODAY}.xlsx')
        save_to_excel(data, excel_file)
    
    elapsed = (time.time() - start) / 60
    log(f"\n총 소요: {elapsed:.1f}분")
    log("=" * 60)

if __name__ == "__main__":
    main()
